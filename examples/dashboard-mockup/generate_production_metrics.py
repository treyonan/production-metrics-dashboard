#!/usr/bin/env python3
"""
generate_production_metrics_v2.py

Reads two mining-site production workbooks and writes
"Production Metrics.html" — a multi-site, Power BI-styled dashboard.

Sites:
  ARQ (Ardmore Quarry)    -> ar Plant Production 25_26.xlsx
  RSQ (Richards Spur Q.)  -> rs Plant Production 2026.xlsx

Data embedded as JSON inside the HTML. Filters, granularity, site
switching, and page navigation are rendered client-side. Chart.js is
inlined into the HTML for offline reliability.

CLI:
  python generate_production_metrics_v2.py \
      --arq <path>.xlsx --rsq <path>.xlsx --out <path>.html \
      [--chartjs <path-to-chart.umd.js>]

Exit codes:
  0 success
  1 source workbook unreadable
  2 unexpected schema (sheet missing)
  3 write failure
"""

from __future__ import annotations
import argparse, datetime as dt, json, os, sys

try:
    from zoneinfo import ZoneInfo
except ImportError:
    ZoneInfo = None  # py<3.9; tz flag will degrade to naive local

# Default IANA timezone for the "Refreshed" label. Mining operations are in
# Oklahoma (Central Time). Override with the --tz CLI flag.
DEFAULT_TZ = "America/Chicago"


def _local_now(tz_name):
    """Return a timezone-aware datetime representing the scheduled-run moment
    in the user's local time. Falls back to naive datetime.now() if zoneinfo
    is unavailable or the tz name is not recognized."""
    if ZoneInfo is not None and tz_name:
        try:
            return dt.datetime.now(ZoneInfo(tz_name))
        except Exception:
            pass
    return dt.datetime.now()


def _format_refresh(t):
    """Human-friendly refresh timestamp for the dashboard header.
    e.g. 'Apr 16, 2026 11:22 PM CT'. Non-zero-padded hour; zero-padded minute.
    %-I isn't portable on Windows, so we compute hour manually."""
    h = t.hour % 12 or 12
    ampm = "AM" if t.hour < 12 else "PM"
    tz_abbr = t.strftime("%Z") or ""
    # Collapse DST/standard pairs to the short common form.
    if tz_abbr in ("CDT", "CST"): tz_abbr = "CT"
    elif tz_abbr in ("EDT", "EST"): tz_abbr = "ET"
    elif tz_abbr in ("MDT", "MST"): tz_abbr = "MT"
    elif tz_abbr in ("PDT", "PST"): tz_abbr = "PT"
    return (t.strftime("%b ") + str(t.day) + t.strftime(", %Y ")
            + f"{h}:{t.minute:02d} {ampm}" + (" " + tz_abbr if tz_abbr else ""))

from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    sys.stderr.write("openpyxl not installed. run: pip install openpyxl\n")
    sys.exit(1)


# ============================================================================
# SITE REGISTRY
# ============================================================================
# Each site has:
#   code, name, accent (hex color), file (xlsx path), circuits (list)
#   Each circuit has:
#     key (unique across all sites),  label, sheet, data_start,
#     availability_col, performance_col, run_time_col, sched_run_col,
#     tons_col (or None if summed from products),
#     product_col (text, or None),
#     downtime_hours_col (numeric, or None),
#     downtime_note_col  (text, or None),
#     truck_cols (dict or None),
#     products (list of {name, tons_col, tph_col?, yield_col?}) or None,
#     stale (bool) -- if True, circuit only appears on DQ page, not in main tabs
#     notes (optional str for tooltips)

ARQ_CIRCUITS = [
    {"key":"arq_primary", "label":"Primary", "sheet":"Primary",
     "data_start":4,
     "availability_col":7, "performance_col":8,
     "run_time_col":6, "sched_run_col":5,
     "tons_col":13, "est_tons_col":11, "rip_rap_pct_col":16,
     "product_col":28,
     "downtime_hours_col":None, "downtime_note_col":None,
     "truck_cols":{"772":36,"773":37,"775":38,"WA600":39,"990":40},
     "products":None, "stale":False,
     "page_config":{
        "kpis":[
          {"l":"Avg Feed TPH",   "calc":"tph", "color":"@accent", "cls":""},
          {"l":"Tons YTD",        "calc":"tons","color":"@accent", "cls":""},
          {"l":"Area Availability","calc":"av", "color":"@accent","cls":"sk"},
          {"l":"Area Performance","calc":"pf", "color":"#107c10","cls":"sp"}]}},

    {"key":"arq_secondary", "label":"Secondary", "sheet":"Secondary",
     "data_start":4,
     "availability_col":35, "performance_col":45,
     "run_time_col":34, "sched_run_col":33,
     "tons_col":41, "est_tons_col":None, "rip_rap_pct_col":None,
     "product_col":4,
     "downtime_hours_col":None, "downtime_note_col":51,
     "truck_cols":None, "products":None, "stale":False,
     "page_config":{
        "kpis":[
          {"l":"Avg Crush TPH",  "calc":"tph", "color":"@accent", "cls":""},
          {"l":"Tons YTD",        "calc":"tons","color":"@accent", "cls":""},
          {"l":"Area Availability","calc":"av", "color":"@accent","cls":"sk"},
          {"l":"Area Performance","calc":"pf", "color":"#107c10","cls":"sp"}]}},

    {"key":"arq_lippman", "label":"Lippman", "sheet":"Lippman Crusher",
     "data_start":4,
     "availability_col":9, "performance_col":10,
     "run_time_col":8, "sched_run_col":7,
     "tons_col":31, "est_tons_col":13, "rip_rap_pct_col":None,
     "product_col":2,
     "downtime_hours_col":None, "downtime_note_col":33,
     "truck_cols":None, "products":None, "stale":False,
     "page_config":{
        "kpis":[
          {"l":"Avg Lippman TPH","calc":"tph", "color":"@accent", "cls":""},
          {"l":"Tons YTD",        "calc":"tons","color":"@accent", "cls":""},
          {"l":"Area Availability","calc":"av", "color":"@accent","cls":"sk"},
          {"l":"Area Performance","calc":"pf", "color":"#107c10","cls":"sp"}]}},

    {"key":"arq_kpi_jci", "label":"KPI JCI", "sheet":"KPI JCI Crusher",
     "data_start":4,
     "availability_col":8, "performance_col":None,
     "run_time_col":7, "sched_run_col":6,
     "tons_col":30, "est_tons_col":10, "rip_rap_pct_col":None,
     "product_col":2,
     "downtime_hours_col":None, "downtime_note_col":33,
     "truck_cols":None, "products":None, "stale":False,
     "page_config":{
        "kpis":[
          {"l":"Avg KPI JCI TPH","calc":"tph", "color":"@accent", "cls":""},
          {"l":"Tons YTD",        "calc":"tons","color":"@accent", "cls":""},
          {"l":"Area Availability","calc":"av", "color":"@accent","cls":"sk"},
          {"l":"Est Tons YTD",    "calc":"est_tons","color":"#6b007b","cls":""}]}},
]

RSQ_CIRCUITS = [
    # Primary: single-product (C2A)
    {"key":"rsq_primary", "label":"Primary", "sheet":"Primary",
     "data_start":2,
     "availability_col":30, "performance_col":31,
     "run_time_col":8, "sched_run_col":7,
     "tons_col":12, "est_tons_col":None, "rip_rap_pct_col":None,
     "product_col":13,
     "downtime_hours_col":29, "downtime_note_col":None,
     "truck_cols":{"775":16,"777":17},
     "products":None, "stale":False,
     "page_config":{
        "kpis":[
          {"l":"Avg C2A Feed TPH","calc":"tph","target":2500,"color":"@accent","cls":""},
          {"l":"C2A Tons YTD",    "calc":"tons","color":"@accent","cls":""},
          {"l":"Area Availability","calc":"av","color":"@accent","cls":"sk"},
          {"l":"Area Performance","calc":"pf","color":"#107c10","cls":"sp"}],
        "tph_target":2500}},

    # Secondary: multi-product
    {"key":"rsq_secondary", "label":"Secondary", "sheet":"Secondary",
     "data_start":2,
     "availability_col":21, "performance_col":22,
     "run_time_col":8, "sched_run_col":7,
     "tons_col":None, "est_tons_col":None, "rip_rap_pct_col":None,
     "product_col":None,
     "downtime_hours_col":20, "downtime_note_col":None,
     "truck_cols":None,
     "products":[
        {"name":"C4",  "tons_col":9,  "tph_col":23},
        {"name":"C8",  "tons_col":11, "tph_col":24, "yield_col":25},
        {"name":"C11", "tons_col":17, "tph_col":26, "yield_col":27},
        {"name":"C12", "tons_col":13, "tph_col":30},
        {"name":"C15", "tons_col":14, "tph_col":28, "yield_col":29},
        {"name":"C60", "tons_col":15},
     ],
     "stale":False,
     "page_config":{
        "kpis":[
          {"l":"Avg C11 Yield",   "calc":"prod_yield","product":"C11","color":"#107c10","cls":""},
          {"l":"C11 Tons YTD",    "calc":"prod_tons", "product":"C11","color":"@accent","cls":""},
          {"l":"Area Availability","calc":"av","color":"@accent","cls":"sk"},
          {"l":"Area Performance","calc":"pf","color":"#107c10","cls":"sp"}]}},

    # Tertiary: full product matrix (crushing availability/performance)
    {"key":"rsq_tertiary", "label":"Tertiary", "sheet":"Tertiary",
     "data_start":2,
     "availability_col":40, "performance_col":41,
     "run_time_col":7, "sched_run_col":6,
     "tons_col":None, "est_tons_col":None, "rip_rap_pct_col":None,
     "product_col":None,
     "downtime_hours_col":None, "downtime_note_col":None,
     "truck_cols":None,
     "feed_avail_col":42,
     "products":[
        {"name":"C17", "tons_col":11, "tph_col":49},
        {"name":"C23", "tons_col":12, "tph_col":50},
        {"name":"C24", "tons_col":13, "tph_col":51},
        {"name":"C25", "tons_col":14, "tph_col":52},
        {"name":"C34", "tons_col":15, "tph_col":68, "yield_col":69},
        {"name":"C40", "tons_col":17, "tph_col":53, "yield_col":55},
        {"name":"C51", "tons_col":19, "tph_col":56, "yield_col":58},
        {"name":"C43", "tons_col":21, "tph_col":22},
        {"name":"C54", "tons_col":24, "tph_col":59, "yield_col":61},
        {"name":"C56", "tons_col":26, "tph_col":62, "yield_col":64},
        {"name":"C58", "tons_col":28, "tph_col":65, "yield_col":67},
     ],
     "stale":False,
     "page_config":{
        "kpis":[
          {"l":"C51 (#57 ODOT) Tons","calc":"prod_tons","product":"C51","goal_pct":0.40,"color":"@accent","cls":""},
          {"l":"C40 Screenings Tons","calc":"prod_tons","product":"C40","goal_pct":0.30,"color":"#6b007b","cls":""},
          {"l":"Area Availability","calc":"av","color":"@accent","cls":"sk"},
          {"l":"Area Performance","calc":"pf","color":"#107c10","cls":"sp"}]}},

    # Sand Plant: RunTime% -> availability, Efficiency -> performance
    {"key":"rsq_sand", "label":"Sand Plant", "sheet":"Sand Plant",
     "data_start":2,
     "availability_col":9, "performance_col":6,
     "run_time_col":8, "sched_run_col":7,
     "tons_col":12, "est_tons_col":None, "rip_rap_pct_col":None,
     "product_col":5,
     "downtime_hours_col":None, "downtime_note_col":None,
     "truck_cols":None, "products":None, "stale":False,
     "page_config":{
        "kpis":[
          {"l":"Total Sand Tons","calc":"tons","color":"@accent","cls":""},
          {"l":"Avg Calc TPH",   "calc":"tph", "color":"@accent","cls":""},
          {"l":"Area Availability","calc":"av","color":"@accent","cls":"sk"},
          {"l":"Area Performance","calc":"pf","color":"#107c10","cls":"sp"}]}},

    # Wash Plant: stale -> DQ only
    {"key":"rsq_wash", "label":"Wash Plant", "sheet":"Wash Plant",
     "data_start":2,
     "availability_col":4, "performance_col":5,
     "run_time_col":3, "sched_run_col":2,
     "tons_col":7, "est_tons_col":None, "rip_rap_pct_col":None,
     "product_col":None,
     "downtime_hours_col":None, "downtime_note_col":None,
     "truck_cols":None, "products":None, "stale":True},

    # Trommel: stale -> DQ only
    {"key":"rsq_trommel", "label":"Trommel", "sheet":"Trommel",
     "data_start":2,
     "availability_col":None, "performance_col":None,
     "run_time_col":4, "sched_run_col":3,
     "tons_col":None, "est_tons_col":None, "rip_rap_pct_col":None,
     "product_col":None,
     "downtime_hours_col":None, "downtime_note_col":None,
     "truck_cols":None, "products":None, "stale":True},
]

SITES = [
    {"code":"ARQ", "name":"Ardmore Quarry",       "accent":"#0078d4",
     "file_arg":"arq", "circuits": ARQ_CIRCUITS},
    {"code":"RSQ", "name":"Richards Spur Quarry", "accent":"#0F5C36",
     "file_arg":"rsq", "circuits": RSQ_CIRCUITS},
]

STALE_HOURS = 48          # source-mtime-based staleness threshold
BLANK_ROW_STOP = 25       # stop extraction after N consecutive blank date cells

# Out-of-range / sanity-check thresholds used during extraction
OOR_PCT_MIN = 0.0
OOR_PCT_MAX = 1.5
OOR_TONS_MAX = 1e6
FUTURE_DATE_DAYS = 7

import re as _re

# ============================================================================
# HEADER-AWARE COLUMN RESOLVER
# ============================================================================
HEADER_SYNONYMS = {
    "availability": [
        "availability", "plant availability", "avail", "run time %",
        "runtime %", "rt %", "avail %", "crushing availability",
        "area availability",
    ],
    "performance": [
        "performance", "plant performance", "perf", "perf %",
        "performance %", "crushing performance", "area performance",
        "performance (by feed scale)", "efficiency",
    ],
    "run_time": [
        "run time", "runtime", "rt", "actual run time", "operating time",
        "op time", "actual runtime", "feed runtime",
    ],
    "sched_run": [
        "sch. run time = shift time", "sch. run time", "scheduled runtime",
        "scheduled run time", "sched run time", "sched. run time",
        "sch runtime", "shift time", "scheduled shift time",
        "scheduled run time = shift time",
    ],
    "tons": [
        "total tons (by j1 belt scale)", "total tons", "tons",
        "tons produced", "production tons", "total tons produced",
        "total tonnage - measured", "tons fed",
        "c2a tons", "secondary tons", "tertiary tons", "sand tons",
        "washed tons", "trommel tons",
    ],
    "est_tons": [
        "est. tons", "est tons", "estimated tons", "est. tons produced",
        "calculated tons fed",
    ],
    "rip_rap_pct": [
        "rip rap %", "rip-rap %", "rip rap percent", "rr %",
        "% rip rap of total primary tons", "% rip rap",
    ],
    "product_label": [
        "product", "product type", "product ran", "product produced",
        "c2a product", "secondary product", "tertiary product",
        "sand product", "trommel product", "a product",
        "note what product ( 1.5\", 2.5\", 3.5\" cr, etc\u2026) or odot",
        "note what product", "note product",
    ],
    "downtime_hrs": [
        "downtime hours", "downtime hrs", "down time hours", "dt hrs",
        "downtime total", "total downtime", "downtime",
    ],
    "downtime_note": [
        "downtime note", "downtime notes", "reason", "downtime reason",
        "comments", "notes", "comments:", "comments regarding down time",
    ],
}

METRIC_COL_MAP = [
    ("availability_col", "availability"),
    ("performance_col", "performance"),
    ("run_time_col", "run_time"),
    ("sched_run_col", "sched_run"),
    ("tons_col", "tons"),
    ("est_tons_col", "est_tons"),
    ("rip_rap_pct_col", "rip_rap_pct"),
    ("product_col", "product_label"),
    ("downtime_hours_col", "downtime_hrs"),
    ("downtime_note_col", "downtime_note"),
]


def _norm_header(s):
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = _re.sub(r"\s+", " ", s)
    return s.rstrip(":")


def _levenshtein(a, b, cap=3):
    if a == b: return 0
    if abs(len(a) - len(b)) > cap: return cap + 1
    if len(a) > len(b): a, b = b, a
    prev = list(range(len(a) + 1))
    for j, cb in enumerate(b, 1):
        curr = [j] + [0] * len(a)
        row_min = curr[0]
        for i, ca in enumerate(a, 1):
            cost = 0 if ca == cb else 1
            curr[i] = min(curr[i - 1] + 1, prev[i] + 1, prev[i - 1] + cost)
            row_min = min(row_min, curr[i])
        if row_min > cap: return cap + 1
        prev = curr
    return prev[-1]


_SYN_EXACT = {}
for _m, _hs in HEADER_SYNONYMS.items():
    for _h in _hs:
        _SYN_EXACT[_norm_header(_h)] = _m


def header_to_metric(hdr):
    k = _norm_header(hdr)
    if not k: return (None, None)
    if k in _SYN_EXACT: return (_SYN_EXACT[k], "exact")
    best = (None, 99)
    for syn, metric in _SYN_EXACT.items():
        cap = 2 if max(len(syn), len(k)) >= 6 else 1
        d = _levenshtein(k, syn, cap=cap)
        if d <= cap and d < best[1]: best = (metric, d)
    if best[0]: return (best[0], "typo")
    return (None, None)


def resolve_circuit_columns(ws, spec):
    header_row = spec.get("header_row") or max(1, spec.get("data_start", 2) - 1)
    max_col = ws.max_column or 1
    matches_by_metric = {m: [] for m in HEADER_SYNONYMS}
    for c in range(1, max_col + 1):
        v = ws.cell(header_row, c).value
        metric, mtype = header_to_metric(v)
        if metric:
            matches_by_metric[metric].append((c, v, mtype))

    per_metric = {}
    warnings = []
    resolved_spec = dict(spec)

    for cfg_key, metric in METRIC_COL_MAP:
        cfg_col = spec.get(cfg_key)
        hits = matches_by_metric.get(metric, [])
        if cfg_col is not None:
            hdr_at_cfg = ws.cell(header_row, cfg_col).value
            metric_at_cfg, match_type = header_to_metric(hdr_at_cfg)
            if metric_at_cfg == metric:
                warn = None
                if match_type == "typo":
                    warn = "%s: header at col %d is %r - looks like a typo of a known header" % (metric, cfg_col, hdr_at_cfg)
                    warnings.append(warn)
                per_metric[metric] = {
                    "configured_col": cfg_col, "resolved_col": cfg_col,
                    "status": "confirmed",
                    "matched_header": str(hdr_at_cfg) if hdr_at_cfg is not None else None,
                    "match_type": match_type, "warning": warn,
                }
            elif hits:
                new_col, new_hdr, new_mt = hits[0]
                warn = ("%s: configured col %d no longer matches (found %r); "
                        "using col %d for %r") % (metric, cfg_col, hdr_at_cfg, new_col, new_hdr)
                warnings.append(warn)
                resolved_spec[cfg_key] = new_col
                per_metric[metric] = {
                    "configured_col": cfg_col, "resolved_col": new_col,
                    "status": "relocated", "matched_header": str(new_hdr),
                    "match_type": new_mt, "warning": warn,
                }
            else:
                warn = ("%s: header at col %d is %r - no synonym match; keeping config col %d"
                        % (metric, cfg_col, hdr_at_cfg, cfg_col))
                warnings.append(warn)
                per_metric[metric] = {
                    "configured_col": cfg_col, "resolved_col": cfg_col,
                    "status": "unverified",
                    "matched_header": str(hdr_at_cfg) if hdr_at_cfg is not None else None,
                    "match_type": None, "warning": warn,
                }
        else:
            if hits:
                new_col, new_hdr, new_mt = hits[0]
                warn = ("%s: header %r at col %d looks like a new source column; "
                        "not enabled (config lists no %s)") % (metric, new_hdr, new_col, cfg_key)
                per_metric[metric] = {
                    "configured_col": None, "resolved_col": None,
                    "status": "detected", "matched_header": str(new_hdr),
                    "match_type": new_mt, "warning": warn,
                }
            else:
                per_metric[metric] = {
                    "configured_col": None, "resolved_col": None,
                    "status": "not_configured", "matched_header": None,
                    "match_type": None, "warning": None,
                }

    return {
        "header_row": header_row,
        "resolved_spec": resolved_spec,
        "per_metric": per_metric,
        "warnings": warnings,
    }


_SHEET_SKIP_SUBSTRINGS = ("piviot", "pivot", "detail", "goals", "sheet1",
                          "sheet2", "sheet3", "chart", "graph", "summary",
                          "lookup", "ref")


def _looks_like_production_sheet(name):
    n = name.strip().lower()
    if not n: return False
    return not any(sub in n for sub in _SHEET_SKIP_SUBSTRINGS)


# ============================================================================
# EXTRACTION
# ============================================================================

def _cell(ws, r: int, c) -> Any:
    return None if c is None else ws.cell(row=r, column=c).value

def _num(v: Any):
    if isinstance(v, bool): return None
    return float(v) if isinstance(v, (int, float)) else None

def _date(v: Any):
    if isinstance(v, dt.datetime): return v.date()
    if isinstance(v, dt.date): return v
    return None

def _text(v: Any):
    if v is None: return None
    s = str(v).strip()
    return s or None


def extract_circuit(ws, spec, sanity=None):
    rows = []
    blank_run = 0
    today = dt.date.today()
    future_cutoff = today + dt.timedelta(days=FUTURE_DATE_DAYS)
    oor = {"av": 0, "pf": 0, "rt": 0, "sr": 0, "tn": 0, "et": 0}
    oor_samples = {"av": [], "pf": [], "tn": []}
    suspect_dates = []
    for r in range(spec.get("data_start", 1), ws.max_row + 1):
        d = _date(_cell(ws, r, 1))
        if d is None:
            blank_run += 1
            if blank_run >= BLANK_ROW_STOP and rows:
                break
            continue
        blank_run = 0

        is_future = d > future_cutoff
        if is_future and len(suspect_dates) < 12:
            suspect_dates.append({"row": r, "date": d.isoformat()})

        av = _num(_cell(ws, r, spec.get("availability_col")))
        pf = _num(_cell(ws, r, spec.get("performance_col")))
        rt = _num(_cell(ws, r, spec.get("run_time_col")))
        sr = _num(_cell(ws, r, spec.get("sched_run_col")))
        tn = _num(_cell(ws, r, spec.get("tons_col")))
        et = _num(_cell(ws, r, spec.get("est_tons_col")))
        rr = _num(_cell(ws, r, spec.get("rip_rap_pct_col")))

        def _pct_ok(v): return v is None or (OOR_PCT_MIN <= v <= OOR_PCT_MAX)
        def _tons_ok(v): return v is None or (0 <= v <= OOR_TONS_MAX)

        if av is not None and not _pct_ok(av):
            oor["av"] += 1
            if len(oor_samples["av"]) < 5: oor_samples["av"].append({"row": r, "value": av})
            av = None
        if pf is not None and not _pct_ok(pf):
            oor["pf"] += 1
            if len(oor_samples["pf"]) < 5: oor_samples["pf"].append({"row": r, "value": pf})
            pf = None
        if rr is not None and not _pct_ok(rr): rr = None
        if tn is not None and not _tons_ok(tn):
            oor["tn"] += 1
            if len(oor_samples["tn"]) < 5: oor_samples["tn"].append({"row": r, "value": tn})
            tn = None
        if et is not None and not _tons_ok(et):
            oor["et"] += 1
            et = None

        row = {
            "d":  d.isoformat(),
            "av": av, "pf": pf, "rt": rt, "sr": sr,
            "tn": tn, "et": et, "rr": rr,
            "pr": _text(_cell(ws, r, spec.get("product_col"))),
            "dh": _num(_cell(ws, r, spec.get("downtime_hours_col"))),
            "dt": _text(_cell(ws, r, spec.get("downtime_note_col"))),
        }
        if is_future: row["_suspect_future_date"] = True

        if spec.get("truck_cols"):
            row["tk"] = {k: (_num(_cell(ws, r, c)) or 0) for k, c in spec["truck_cols"].items()}

        if spec.get("products"):
            pm = {}
            tn_sum = 0.0
            any_tons = False
            for p in spec["products"]:
                t = _num(_cell(ws, r, p["tons_col"]))
                tph = _num(_cell(ws, r, p.get("tph_col"))) if p.get("tph_col") else None
                yld = _num(_cell(ws, r, p.get("yield_col"))) if p.get("yield_col") else None
                if t is not None and not _tons_ok(t):
                    oor["tn"] += 1
                    t = None
                pm[p["name"]] = {"tn": t, "tph": tph, "yld": yld}
                if t is not None:
                    tn_sum += t
                    any_tons = True
            row["pm"] = pm
            if spec.get("tons_col") is None and any_tons:
                row["tn"] = tn_sum

        populated = any(row.get(k) not in (None, 0) for k in ("av","pf","rt","tn","et"))
        row["_p"] = populated
        rows.append(row)

    if sanity is not None:
        sanity["oor_counts"] = oor
        sanity["oor_samples"] = oor_samples
        sanity["suspect_future_dates"] = suspect_dates
        sanity["rows_total"] = len(rows)
        sanity["rows_populated"] = sum(1 for r in rows if r.get("_p"))
    return rows


_GOALS_HDR_SYNONYMS = {
    "size":   ["size", "product size", "screen size"],
    "code":   ["code", "product code", "sku"],
    "pct":    ["%", "percent", "goal %", "target %", "share", "goal"],
    "tons":   ["goal tons", "est tons", "est. tons", "target tons", "daily tons", "tons"],
    "target": ["daily target", "daily target tons", "target"],
}


def _goals_find_columns(g):
    found = {"size_col": None, "code_col": None, "pct_col": None,
             "tons_col": None, "header_row": None}
    max_r = min(g.max_row or 1, 10)
    max_c = min(g.max_column or 1, 30)
    syn_to_field = {}
    for field, syns in _GOALS_HDR_SYNONYMS.items():
        for s in syns: syn_to_field[_norm_header(s)] = field
    for r in range(1, max_r + 1):
        hits = {}
        for c in range(1, max_c + 1):
            v = _norm_header(g.cell(r, c).value)
            if not v: continue
            if v in syn_to_field:
                field = syn_to_field[v]
                hits.setdefault(field, c)
        if "code" in hits and "pct" in hits:
            found["header_row"] = r
            found["size_col"] = hits.get("size")
            found["code_col"] = hits["code"]
            found["pct_col"] = hits["pct"]
            found["tons_col"] = hits.get("tons")
            return found
    return found


def read_rsq_goals(wb):
    out = {"daily_target": None, "products": {},
           "resolution": {"header_row": None, "by": None, "warning": None}}
    if "Goals" not in wb.sheetnames:
        out["resolution"]["warning"] = "no 'Goals' sheet found in workbook"
        return out
    g = wb["Goals"]
    scan = _goals_find_columns(g)
    if scan["code_col"] and scan["pct_col"]:
        out["resolution"]["by"] = "header"
        out["resolution"]["header_row"] = scan["header_row"]
        size_c, code_c = scan["size_col"], scan["code_col"]
        pct_c, tons_c = scan["pct_col"], scan["tons_col"]
        start = (scan["header_row"] or 0) + 1
    else:
        out["resolution"]["by"] = "fallback"
        out["resolution"]["warning"] = "Goals sheet has no recognized header row; using legacy layout"
        size_c, code_c, pct_c, tons_c, start = 6, 7, 8, 9, 1

    target_val = None
    for r in range(1, min(g.max_row + 1, 20)):
        for c in range(1, min(g.max_column + 1, 30)):
            v = _norm_header(g.cell(r, c).value)
            if v in ("daily target", "daily target tons"):
                cand = g.cell(r, c + 1).value
                if isinstance(cand, (int, float)):
                    target_val = float(cand); break
        if target_val is not None: break
    if target_val is None:
        try:
            dt_val = g.cell(4, 8).value
            if isinstance(dt_val, (int, float)): target_val = float(dt_val)
        except Exception: pass
    out["daily_target"] = target_val

    for r in range(start, min(g.max_row + 1, 100)):
        size = g.cell(r, size_c).value if size_c else None
        code = g.cell(r, code_c).value if code_c else None
        pct = g.cell(r, pct_c).value if pct_c else None
        tons = g.cell(r, tons_c).value if tons_c else None
        if isinstance(code, str) and isinstance(pct, (int, float)) and 0 < pct <= 1:
            out["products"][code.strip()] = {
                "pct": float(pct),
                "tons": float(tons) if isinstance(tons, (int, float)) else None,
                "size": str(size) if size is not None else None,
            }
    return out


def build_site(site, file_path):
    if not file_path.exists():
        sys.stderr.write(f"{site['code']} source not found: {file_path}\n")
        sys.exit(1)
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
    except Exception as e:
        sys.stderr.write(f"{site['code']} cannot open: {e}\n"); sys.exit(1)

    mtime = dt.datetime.fromtimestamp(file_path.stat().st_mtime)
    now = dt.datetime.now()
    stale_file = (now - mtime).total_seconds() > STALE_HOURS * 3600

    if site["code"] == "RSQ":
        goals = read_rsq_goals(wb)
    else:
        goals = {"daily_target": None, "products": {},
                 "resolution": {"header_row": None, "by": None, "warning": None}}

    circuits = {}
    order_active = []
    order_stale = []
    config_health = {}
    site_warnings = []
    consumed_sheets = set()

    for spec in site["circuits"]:
        if spec["sheet"] not in wb.sheetnames:
            sys.stderr.write(f"{site['code']} sheet missing: {spec['sheet']}\n")
            sys.exit(2)
        consumed_sheets.add(spec["sheet"])
        ws = wb[spec["sheet"]]

        resolution = resolve_circuit_columns(ws, spec)
        resolved_spec = resolution["resolved_spec"]

        sanity = {}
        rs = extract_circuit(ws, resolved_spec, sanity=sanity)

        pcfg = spec.get("page_config")
        if pcfg and goals["products"]:
            pcfg = json.loads(json.dumps(pcfg))
            for kp in pcfg.get("kpis", []):
                prod = kp.get("product")
                if prod and prod in goals["products"]:
                    gp = goals["products"][prod]["pct"]
                    kp["goal_pct"] = gp
                    kp["goal_tons"] = goals["products"][prod].get("tons")

        has_av_data = any(isinstance(r.get("av"), (int, float)) for r in rs)
        has_pf_data = any(isinstance(r.get("pf"), (int, float)) for r in rs)
        circuits[spec["key"]] = {
            "label": spec["label"],
            "sheet": spec["sheet"],
            "has_performance": (resolved_spec.get("performance_col") is not None) and has_pf_data,
            "has_availability": (resolved_spec.get("availability_col") is not None) and has_av_data,
            "has_downtime_hours": resolved_spec.get("downtime_hours_col") is not None,
            "has_downtime_note": resolved_spec.get("downtime_note_col") is not None,
            "has_trucks": bool(resolved_spec.get("truck_cols")),
            "truck_labels": list(resolved_spec["truck_cols"].keys()) if resolved_spec.get("truck_cols") else [],
            "has_products": bool(resolved_spec.get("products")),
            "product_names": [p["name"] for p in resolved_spec["products"]] if resolved_spec.get("products") else [],
            "stale": spec.get("stale", False),
            "page_config": pcfg,
            "rows": rs,
        }
        (order_stale if spec.get("stale") else order_active).append(spec["key"])

        status_tally = {"confirmed": 0, "relocated": 0, "unverified": 0,
                        "detected": 0, "not_configured": 0}
        warn_list = []
        for mname, info in resolution["per_metric"].items():
            status_tally[info["status"]] += 1
            if info.get("warning"):
                warn_list.append({"metric": mname, "status": info["status"],
                                  "message": info["warning"]})
        config_health[spec["key"]] = {
            "label": spec["label"],
            "sheet": spec["sheet"],
            "header_row": resolution["header_row"],
            "per_metric": resolution["per_metric"],
            "status_tally": status_tally,
            "warnings": warn_list,
            "sanity": sanity,
        }

    unregistered = []
    for name in wb.sheetnames:
        if name in consumed_sheets: continue
        if _looks_like_production_sheet(name):
            ws_u = wb[name]
            found_metrics = set()
            for c in range(1, min(ws_u.max_column or 1, 30) + 1):
                for r in range(1, min(ws_u.max_row or 1, 5) + 1):
                    metric, _ = header_to_metric(ws_u.cell(r, c).value)
                    if metric: found_metrics.add(metric)
            if found_metrics:
                unregistered.append({
                    "sheet": name,
                    "detected_metrics": sorted(found_metrics),
                    "max_row": ws_u.max_row,
                    "max_col": ws_u.max_column,
                })
                site_warnings.append(
                    f"sheet '{name}' has production-like headers "
                    f"({', '.join(sorted(found_metrics))}) but is not mapped"
                )

    return {
        "code": site["code"],
        "name": site["name"],
        "accent": site["accent"],
        "file": str(file_path),
        "file_mtime": mtime.isoformat(timespec="seconds"),
        "stale_file": stale_file,
        "circuits": circuits,
        "circuit_order_active": order_active,
        "circuit_order_stale": order_stale,
        "goals": goals,
        "config_health": config_health,
        "unregistered_sheets": unregistered,
        "site_warnings": site_warnings,
        "all_sheets": list(wb.sheetnames),
    }


# ============================================================================
# RENDER
# ============================================================================

def load_chartjs(path: Path) -> str:
    try:
        return path.read_text(encoding="utf-8")
    except Exception as e:
        sys.stderr.write(f"warn: cannot load Chart.js bundle: {e}\n")
        return ""


def render_html(payload: dict, chartjs: str) -> str:
    data_json = json.dumps(payload, separators=(",", ":"), default=str)
    html_out = TEMPLATE
    html_out = html_out.replace("__CHARTJS_INLINE__", chartjs)
    html_out = html_out.replace("__DATA_JSON__", data_json)
    return html_out


def _fingerprint_site(site_data):
    fp = {"code": site_data["code"], "sheets": list(site_data.get("all_sheets", [])),
          "circuits": {}}
    for ckey, ch in site_data.get("config_health", {}).items():
        cols = {}
        for m, info in ch["per_metric"].items():
            cols[m] = {"col": info["resolved_col"],
                        "hdr": info["matched_header"],
                        "st": info["status"]}
        fp["circuits"][ckey] = {
            "sheet": ch["sheet"],
            "header_row": ch["header_row"],
            "cols": cols,
        }
    return fp


def _diff_snapshots(prev, curr):
    drift = []
    if not prev: return drift
    prev_sites = {s["code"]: s for s in prev.get("sites", [])}
    for s in curr.get("sites", []):
        code = s["code"]
        p = prev_sites.get(code)
        if not p:
            drift.append({"site": code, "kind": "new_site"}); continue
        added = set(s["sheets"]) - set(p.get("sheets", []))
        removed = set(p.get("sheets", [])) - set(s["sheets"])
        for n in sorted(added): drift.append({"site": code, "kind": "sheet_added", "sheet": n})
        for n in sorted(removed): drift.append({"site": code, "kind": "sheet_removed", "sheet": n})
        for ckey, cur_c in s["circuits"].items():
            prev_c = p["circuits"].get(ckey)
            if not prev_c:
                drift.append({"site": code, "kind": "circuit_added", "circuit": ckey}); continue
            for m, cur_col in cur_c["cols"].items():
                prev_col = prev_c["cols"].get(m)
                if not prev_col: continue
                if cur_col["col"] != prev_col["col"]:
                    drift.append({"site": code, "kind": "col_moved",
                                  "circuit": ckey, "metric": m,
                                  "from": prev_col["col"], "to": cur_col["col"]})
                elif cur_col["hdr"] != prev_col["hdr"]:
                    drift.append({"site": code, "kind": "header_renamed",
                                  "circuit": ckey, "metric": m,
                                  "from": prev_col["hdr"], "to": cur_col["hdr"]})
    return drift


def build(arq_path, rsq_path, out, chartjs_path, tz_name=DEFAULT_TZ):
    now = _local_now(tz_name)
    today = now.date()

    site_data = []
    for site in SITES:
        src_path = arq_path if site["code"] == "ARQ" else rsq_path
        site_data.append(build_site(site, src_path))

    snap_path = out.parent / "schema_snapshot.json"
    prev_snap = None
    try:
        if snap_path.exists():
            prev_snap = json.loads(snap_path.read_text(encoding="utf-8"))
    except Exception:
        prev_snap = None

    curr_snap = {
        "generated_at": now.isoformat(timespec="seconds"),
        "sites": [_fingerprint_site(sd) for sd in site_data],
    }
    drift = _diff_snapshots(prev_snap, curr_snap)
    if prev_snap:
        curr_snap["prev_generated_at"] = prev_snap.get("generated_at")

    try:
        snap_path.write_text(json.dumps(curr_snap, indent=2, default=str),
                             encoding="utf-8")
    except Exception as e:
        sys.stderr.write(f"warn: schema_snapshot write failed: {e}\n")

    payload = {
        "as_of": today.isoformat(),
        "generated_at": _format_refresh(now),
        "generated_at_iso": now.isoformat(timespec="seconds"),
        "stale_hours": STALE_HOURS,
        "sites": site_data,
        "site_order": [s["code"] for s in SITES],
        "schema_drift": drift,
        "schema_snapshot_prev": prev_snap.get("generated_at") if prev_snap else None,
    }

    chartjs = load_chartjs(chartjs_path) if chartjs_path else ""
    html_out = render_html(payload, chartjs)

    try:
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(html_out, encoding="utf-8")
    except Exception as e:
        sys.stderr.write(f"write failed: {e}\n"); sys.exit(3)


# ============================================================================
# HTML / CSS / JS TEMPLATE
# ============================================================================

TEMPLATE = r"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<title>Dolese Production Metrics</title>
<meta name="viewport" content="width=device-width,initial-scale=1">
<style>
*{box-sizing:border-box;margin:0;padding:0;}
body{background:#201f1e;font-family:'Segoe UI',system-ui,sans-serif;}
.shell{background:#201f1e;overflow:hidden;}
.topbar{height:40px;background:#201f1e;border-bottom:1px solid #484644;display:flex;align-items:center;padding:0 10px;}
.tlogo{display:flex;align-items:center;gap:7px;padding-right:12px;border-right:1px solid #484644;}
.tlogo b{color:#f3f2f1;font-size:12px;}
.tlogo s{color:#a19f9d;font-size:11px;text-decoration:none;}
.tbtn{height:40px;padding:0 11px;background:none;border:none;border-right:1px solid #484644;color:#c8c6c4;font-size:11px;cursor:pointer;font-family:'Segoe UI',sans-serif;}
.tbtn:hover{background:#484644;color:#f3f2f1;}
.site-tog{display:flex;align-items:center;border-right:1px solid #484644;}
.stbtn{height:40px;padding:0 14px;background:none;border:none;font-size:11px;cursor:pointer;font-family:'Segoe UI',sans-serif;font-weight:600;color:#c8c6c4;border-bottom:3px solid transparent;transition:color .15s,border-color .15s,background .15s;}
.stbtn:hover{background:#3b3a39;color:#f3f2f1;}
.stbtn.on{color:#f3f2f1;}
.canvas{display:flex;background:#f3f2f1;min-height:calc(100vh - 71px);}
.sidebar{width:180px;flex-shrink:0;background:#fafaf9;border-right:1px solid #e1dfdd;padding:7px 5px;display:flex;flex-direction:column;gap:5px;}
.main{flex:1;padding:7px;display:flex;flex-direction:column;gap:6px;min-width:0;overflow-x:hidden;}
.sb{background:#fff;border:1px solid #e1dfdd;border-radius:2px;}
.sh{font-size:10px;font-weight:700;color:#201f1e;padding:5px 7px 4px;background:#f3f2f1;border-bottom:1px solid #e1dfdd;letter-spacing:.04em;}
.sd{padding:4px 5px;}
.ss{font-size:9px;font-weight:700;color:#605e5c;letter-spacing:.04em;padding:5px 2px 2px;}
.sr{display:flex;align-items:center;gap:3px;padding:3px 2px;}
.sr label{font-size:10px;color:#201f1e;font-weight:600;width:30px;flex-shrink:0;}
.sr select{flex:1;font-size:11px;height:24px;border:1px solid #605e5c;border-radius:2px;padding:0 3px;background:#fff;color:#201f1e;font-family:'Segoe UI',sans-serif;cursor:pointer;}
.gbr{display:flex;gap:2px;flex-wrap:wrap;padding:3px 2px;}
.gb{padding:3px 5px;font-size:10px;border:1px solid #605e5c;border-radius:2px;background:#fff;color:#201f1e;cursor:pointer;font-weight:600;font-family:'Segoe UI',sans-serif;}
.gb.on{background:var(--accent,#0078d4);color:#fff;border-color:var(--accent,#0078d4);}
.gb:hover:not(.on){background:#f3f2f1;}
.pnv{display:flex;align-items:center;gap:3px;padding:4px 2px 3px;}
.pnb{width:21px;height:21px;background:#fff;border:1px solid #605e5c;border-radius:2px;cursor:pointer;font-size:11px;color:#201f1e;font-family:'Segoe UI',sans-serif;}
.pnb:hover{background:#f3f2f1;}
.pnl{flex:1;text-align:center;font-size:11px;font-weight:700;color:#201f1e;}
.si{display:flex;align-items:center;gap:6px;padding:3px 2px;cursor:pointer;border-radius:2px;}
.si:hover{background:#f3f2f1;}
.srb{width:12px;height:12px;border:1.5px solid #605e5c;border-radius:50%;flex-shrink:0;position:relative;}
.srb.on{border-color:var(--accent,#0078d4);}
.srb.on::after{content:'';position:absolute;width:6px;height:6px;background:var(--accent,#0078d4);border-radius:50%;top:1.5px;left:1.5px;}
.sl{font-size:11px;color:#201f1e;}
.kg{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:6px;}
.kc{background:#fff;border:1px solid #e1dfdd;border-radius:2px;padding:8px 10px 6px;border-top:3px solid var(--accent,#0078d4);}
.kc.sk{border-top-color:var(--accent,#0078d4);}
.kc.sp{border-top-color:#107c10;}
.kc.so{border-top-color:#e66c37;}
.kc.sr{border-top-color:#a4262c;}
.kc.off{border-top-color:#8a8886;opacity:.65;}
.kl{font-size:10px;color:#605e5c;margin-bottom:2px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;}
.kv{font-size:20px;font-weight:300;color:#201f1e;line-height:1.1;}
.km{font-size:10px;margin-top:2px;}
.kb{height:2px;background:#e1dfdd;border-radius:1px;margin-top:5px;}
.kf{height:2px;border-radius:1px;}
.vi{background:#fff;border:1px solid #e1dfdd;border-radius:2px;overflow:hidden;}
.vh{display:flex;align-items:center;padding:5px 7px 2px;}
.vt{font-size:11px;font-weight:700;color:#201f1e;flex:1;}
.vn{font-size:9px;color:#a19f9d;padding:0 7px 2px;font-style:italic;}
.vb{padding:3px 7px 6px;}
.cw{position:relative;height:175px;}
.r2{display:grid;gap:6px;grid-template-columns:3fr 2fr;}
.r3{display:grid;gap:6px;grid-template-columns:2fr 3fr;}
.lg{display:flex;gap:6px;flex-wrap:wrap;padding:0 7px 3px;}
.li{display:flex;align-items:center;gap:3px;font-size:10px;color:#323130;}
.ls{width:9px;height:9px;border-radius:1px;flex-shrink:0;}
.chip{display:inline-flex;background:#dce8fa;color:#004578;font-size:10px;padding:2px 7px;border-radius:10px;font-weight:600;}
.chip.w{background:#fff4ce;color:#7a4800;}
.chip.r{background:#fde7e9;color:#a4262c;}
.chip.g{background:#dff6dd;color:#0e5c0e;}
.chip.site{background:var(--accent-tint,#e6f2fb);color:var(--accent,#0078d4);}
.mxw{overflow-x:auto;}
.mx{width:100%;border-collapse:collapse;font-size:10px;}
.mx th{background:#f3f2f1;font-weight:700;color:#201f1e;padding:4px 7px;text-align:right;border-bottom:1px solid #e1dfdd;white-space:nowrap;}
.mx th:first-child{text-align:left;}
.mx td{padding:3px 7px;text-align:right;color:#323130;border-bottom:.5px solid #f3f2f1;white-space:nowrap;}
.mx td:first-child{text-align:left;color:#605e5c;font-weight:600;}
.mx tr:hover td{background:#fafaf9;}
.mxt td{background:#f3f2f1;font-weight:700;border-top:1px solid #c8c6c4;color:#201f1e;}
.ptabs{background:#201f1e;height:31px;display:flex;align-items:flex-end;padding:0 7px;gap:1px;border-top:1px solid #484644;overflow-x:auto;}
.pt{height:27px;padding:0 10px;font-size:11px;cursor:pointer;border-radius:2px 2px 0 0;background:#3b3a39;color:#c8c6c4;border:none;font-weight:600;font-family:'Segoe UI',sans-serif;white-space:nowrap;}
.pt.on{background:#f3f2f1;color:#201f1e;}
.pt:hover:not(.on){background:#484644;color:#f3f2f1;}
.warn-bar{background:#fff4ce;border:1px solid #f2c811;border-radius:2px;padding:6px 10px;font-size:10px;color:#7a4800;display:flex;gap:6px;align-items:center;}
.warn-bar.err{background:#fde7e9;border-color:#a4262c;color:#a4262c;margin-top:4px;}
.site-strip{padding:5px 7px;border-bottom:2px solid var(--accent,#0078d4);background:var(--accent-tint,#e6f2fb);display:flex;align-items:center;gap:8px;}
.site-id{font-size:13px;font-weight:700;color:var(--accent,#0078d4);}
.site-nm{font-size:10px;color:#605e5c;}
.nd{display:flex;flex-direction:column;align-items:center;justify-content:center;height:150px;gap:5px;color:#605e5c;font-size:11px;}
.bn{display:flex;gap:3px;flex-wrap:wrap;}
.bb{padding:4px 10px;font-size:11px;border:1px solid #605e5c;border-radius:2px;background:#fff;color:#201f1e;cursor:pointer;font-weight:600;font-family:'Segoe UI',sans-serif;}
.bb.on{background:var(--accent,#0078d4);color:#fff;border-color:var(--accent,#0078d4);}
.bb:hover:not(.on){background:#f3f2f1;}
.oee-lbl{display:inline-flex;padding:1px 5px;border-radius:2px;font-size:9px;font-weight:700;}
.av-l{background:#dce8fa;color:#004578;}
.pf-l{background:#dff6dd;color:#0e5c0e;}
</style>
</head><body>
<div class="shell">
<div class="topbar">
  <div class="tlogo">
    <svg width="14" height="14" viewBox="0 0 14 14"><path d="M1 2 L10 7 L1 12 L4 7 Z" fill="#F2C811"/><path d="M1 2 L10 7 L1 12 L4 7 Z" fill="#0F5C36" transform="translate(1,1) scale(.85)"/></svg>
    <b>Production Metrics</b><s>&nbsp;· Dolese Bros Co</s>
  </div>
  <div class="site-tog" id="site-tog"></div>
  <button class="tbtn" onclick="window.print()">Print</button>
  <div style="flex:1"></div>
  <span style="font-size:10px;color:#a19f9d;padding:0 10px;font-style:italic" id="refresh-lbl"></span>
</div>
<div class="canvas" id="canvas">
<div class="sidebar">
  <div class="site-strip" id="site-strip"></div>
  <div class="sb">
    <div class="sh">TIME FILTER</div>
    <div class="sd">
      <div class="ss">Granularity</div>
      <div class="gbr" id="gbr"></div>
      <div class="ss" style="margin-top:4px">Date hierarchy</div>
      <div class="sr"><label>Year</label><select id="yr-sl"></select></div>
      <div class="sr"><label>Qtr</label><select id="qt-sl"></select></div>
      <div class="sr"><label>Month</label><select id="mo-sl"></select></div>
      <div class="ss" style="margin-top:4px">Period navigator</div>
      <div class="pnv"><button class="pnb" onclick="stepP(-1)">&#9664;</button><span class="pnl" id="pnl">All</span><button class="pnb" onclick="stepP(1)">&#9654;</button></div>
    </div>
  </div>
  <div class="sb"><div class="sh">CIRCUIT FOCUS</div><div class="sd" id="circ-sl"></div></div>
  <div style="flex:1"></div>
  <div class="sb"><div class="sh">ACTIVE FILTERS</div><div class="sd" id="chips" style="display:flex;flex-wrap:wrap;gap:3px;min-height:20px"></div></div>
  <div class="sb"><div class="sh">SOURCE</div>
    <div class="sd" style="font-size:10px;color:#605e5c;line-height:1.35">
      <div id="src-file">&mdash;</div>
      <div style="margin-top:4px">Modified: <b id="src-mtime">&mdash;</b></div>
      <div>As of: <b id="src-asof">&mdash;</b></div>
    </div>
  </div>
</div>
<div class="main">
  <div style="display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:4px">
    <div>
      <div style="font-size:9px;color:#605e5c;margin-bottom:3px;font-weight:700;letter-spacing:.04em">BOOKMARK NAVIGATOR</div>
      <div class="bn" id="bnv"></div>
    </div>
    <span class="chip site" id="xfchip" style="font-size:10px;padding:3px 8px;border-radius:2px;font-weight:700">Cross-filter &middot; Period-linked</span>
  </div>
  <div id="wbar"></div>
  <div id="kg-w"></div>
  <div class="r2" id="r2"></div>
  <div class="r3" id="r3"></div>
  <div class="vi">
    <div class="vh"><div class="vt" id="mxt">Matrix</div></div>
    <div class="vn">Aggregated from daily records</div>
    <div class="vb"><div class="mxw" id="mxb"></div></div>
  </div>
</div>
</div>
<div class="ptabs" id="ptabs"></div>
</div>

<script>
// --- inlined Chart.js 4.4.1 (UMD) -----------------------------------------
__CHARTJS_INLINE__
// --------------------------------------------------------------------------
</script>
<script>
var DATA = __DATA_JSON__;

var MO = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];
var QM = {Q1:[0,1,2],Q2:[3,4,5],Q3:[6,7,8],Q4:[9,10,11]};
var GRAINS = ['Hourly','Daily','Weekly','Monthly','Quarterly','Annually'];
var RAW_GRAINS = {Hourly:1};  // grains that need shift-level data we don't have

// build {code: siteData} map
var SITES_BY_CODE = {};
DATA.sites.forEach(function(s){ SITES_BY_CODE[s.code] = s; });

// Per-site circuit color palette (5 distinct hues) — generated around accent
var CIRC_PALETTES = {};
DATA.sites.forEach(function(s){
  // simple fixed palette per site, anchored by accent then 4 contrasting tones
  CIRC_PALETTES[s.code] = [s.accent, '#107c10', '#e66c37', '#6b007b', '#00b294', '#a4262c'];
});

// Standard PAGES: Overview, OEE, (dynamic circuits), DQ
function pagesForSite(code){
  var s = SITES_BY_CODE[code];
  var pages = [{k:'overview',label:'Overview'}, {k:'oee',label:'OEE'}];
  s.circuit_order_active.forEach(function(k){
    pages.push({k:k, label:s.circuits[k].label});
  });
  pages.push({k:'dq', label:'Data Quality'});
  return pages;
}

// Pre-parse dates across every site+circuit
DATA.sites.forEach(function(s){
  Object.keys(s.circuits).forEach(function(k){
    s.circuits[k].rows.forEach(function(r){
      r._d = new Date(r.d + 'T00:00:00');
      r._y = r._d.getFullYear();
      r._m = r._d.getMonth();
      r._q = 'Q' + (Math.floor(r._m/3)+1);
    });
  });
});

function allYearsForSite(code){
  var s = SITES_BY_CODE[code];
  var ys = new Set();
  Object.keys(s.circuits).forEach(function(k){
    s.circuits[k].rows.forEach(function(r){ ys.add(r._y); });
  });
  return Array.from(ys).sort();
}

// Global app state
var ST = {
  site: DATA.site_order[0],
  page: 'overview',
  grain: 'Monthly',
  yr: null,
  qt: 'All',
  mo: 'All',
  pidx: 0,
  circ: null   // currently-focused circuit (within site)
};
// Initialize ST.yr and ST.circ from the active site
(function(){
  var yrs = allYearsForSite(ST.site);
  var cur = new Date(DATA.as_of).getFullYear();
  ST.yr = yrs.indexOf(cur) >= 0 ? cur : (yrs[yrs.length-1] || 'All');
  ST.circ = SITES_BY_CODE[ST.site].circuit_order_active[0];
})();

var CH = {};

// ---- Formatters ----
function fK(n){if(n==null||isNaN(n))return '\u2014';if(n>=1e6)return(n/1e6).toFixed(2)+'M';if(n>=1000)return(n/1000).toFixed(1)+'k';return Math.round(n).toLocaleString();}
function fN(n,d){if(n==null||isNaN(n))return '\u2014';return d?n.toFixed(d):Math.round(n).toLocaleString();}
function fP(n){if(n==null||isNaN(n))return '\u2014';return(n*100).toFixed(1)+'%';}
function avg(a){var v=a.filter(function(x){return x!=null&&!isNaN(x);});return v.length?v.reduce(function(s,x){return s+x;},0)/v.length:null;}
function sum(a){var v=a.filter(function(x){return x!=null&&!isNaN(x);});return v.length?v.reduce(function(s,x){return s+x;},0):null;}
function isoWeek(d){var t=new Date(Date.UTC(d.getFullYear(),d.getMonth(),d.getDate()));var day=t.getUTCDay()||7;t.setUTCDate(t.getUTCDate()+4-day);var y=new Date(Date.UTC(t.getUTCFullYear(),0,1));return Math.ceil((((t-y)/86400000)+1)/7);}
function escH(s){s=String(s==null?'':s); return s.replace(/[&<>\"']/g,function(c){return {'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c];});}
function hexToRgba(h,a){h=h.replace('#',''); var r=parseInt(h.slice(0,2),16),g=parseInt(h.slice(2,4),16),b=parseInt(h.slice(4,6),16); return 'rgba('+r+','+g+','+b+','+a+')';}

function kc(id){if(CH[id]){try{CH[id].destroy();}catch(e){}delete CH[id];}}
function cr(){Object.keys(CH).forEach(function(id){kc(id);});document.getElementById('r2').innerHTML='';document.getElementById('r3').innerHTML='';}
function isRawGrain(){return !!RAW_GRAINS[ST.grain];}
function rawPlaceholder(){
  var nd = '<div class="vi"><div class="vh"><div class="vt">'+escH(ST.grain)+' granularity \u2014 shift-level source required</div></div>'+
           '<div class="vn">Connect raw shift log data to unlock '+escH(ST.grain.toLowerCase())+'-level charts</div>'+
           '<div class="vb"><div class="nd"><div style="font-size:20px;color:#f2c811">\u26A0</div>'+
           '<div>Daily workbooks do not carry '+escH(ST.grain.toLowerCase())+' detail</div></div></div></div>';
  document.getElementById('r2').innerHTML = nd + nd;
  document.getElementById('r3').innerHTML = nd + nd;
}

// ---- Site / page helpers ----
function curSite(){ return SITES_BY_CODE[ST.site]; }
function curCirc(){ return curSite().circuits[ST.circ]; }
function circColor(k){
  var s=curSite();
  var idx = s.circuit_order_active.indexOf(k);
  if(idx<0) idx = s.circuit_order_stale.indexOf(k);
  return CIRC_PALETTES[s.code][(idx<0?0:idx) % CIRC_PALETTES[s.code].length];
}

// ---- Filtering ----
function filterRows(rows){
  return rows.filter(function(r){
    if(ST.yr!=='All' && r._y!==ST.yr) return false;
    if(ST.qt!=='All'){var mons=QM[ST.qt]||[]; if(mons.indexOf(r._m)<0) return false;}
    if(ST.mo!=='All'){var mi=MO.indexOf(ST.mo); if(r._m!==mi) return false;}
    return true;
  });
}

// ---- Granularity aggregation ----
function aggregate(rows, field, op){
  var groups = {};
  var order = [];
  rows.forEach(function(r){
    var key;
    if(ST.grain==='Daily') key=r.d;
    else if(ST.grain==='Weekly') key=r._y+'-W'+String(isoWeek(r._d)).padStart(2,'0');
    else if(ST.grain==='Monthly') key=r._y+'-'+String(r._m+1).padStart(2,'0');
    else if(ST.grain==='Quarterly') key=r._y+'-'+r._q;
    else key=String(r._y);
    if(groups[key]===undefined){groups[key]=[];order.push(key);}
    if(r[field]!=null) groups[key].push(r[field]);
  });
  order.sort();
  var labels = order.map(function(k){
    if(ST.grain==='Monthly'){var p=k.split('-');return MO[parseInt(p[1])-1]+(ST.yr==='All'?' '+p[0].slice(2):'');}
    if(ST.grain==='Quarterly'){var p=k.split('-');return p[1]+(ST.yr==='All'?' '+p[0].slice(2):'');}
    if(ST.grain==='Weekly'){var p=k.split('-W');return 'W'+p[1]+(ST.yr==='All'?" '"+p[0].slice(2):'');}
    if(ST.grain==='Daily'){var d=new Date(k+'T00:00:00'); return (d.getMonth()+1)+'/'+d.getDate();}
    return k;
  });
  var values = order.map(function(k){
    var vs = groups[k];
    if(!vs.length) return null;
    if(op==='avg') return vs.reduce(function(s,x){return s+x;},0)/vs.length;
    return vs.reduce(function(s,x){return s+x;},0);
  });
  return {labels:labels, values:values, keys:order};
}

// ---- KPI helpers ----
function kpiTile(l, v, m, c, p, cls){
  return '<div class="kc '+(cls||'')+'"><div class="kl">'+l+'</div><div class="kv">'+v+'</div><div class="km" style="color:'+c+'">'+m+'</div><div class="kb"><div class="kf" style="width:'+Math.max(0,Math.min(100,p))+'%;background:'+c+'"></div></div></div>';
}
function circMetrics(key){
  var c = curSite().circuits[key];
  var all = filterRows(c.rows);
  var av = avg(all.map(function(r){return r.av;}));
  var pf = avg(all.map(function(r){return r.pf;}));
  var tn = sum(all.map(function(r){return r.tn;}));
  var et = sum(all.map(function(r){return r.et;}));
  var rt = sum(all.map(function(r){return r.rt;}));
  var sr = sum(all.map(function(r){return r.sr;}));
  var dh = sum(all.map(function(r){return r.dh;}));
  // Derived TPH = total tons / total run hours (site-typical calculation)
  var tph = (tn!=null&&rt!=null&&rt>0) ? tn/rt : null;
  return {av:av, pf:pf, tn:tn, et:et, rt:rt, sr:sr, dh:dh, tph:tph, util:(rt!=null&&sr!=null&&sr>0?rt/sr:null), n:all.length, _rows:all};
}

function productAgg(rows, prodName){
  // Aggregate product-level tons and avg yield across filtered rows
  var tn=0, tnAny=false, yldVals=[];
  rows.forEach(function(r){
    if(!r.pm || !r.pm[prodName]) return;
    var p = r.pm[prodName];
    if(p.tn!=null){ tn += p.tn; tnAny=true; }
    if(p.yld!=null) yldVals.push(p.yld);
  });
  return {
    tn: tnAny ? tn : null,
    yld: yldVals.length ? yldVals.reduce(function(s,x){return s+x;},0)/yldVals.length : null
  };
}

function configuredKpis(c, m, s, cColor){
  var cfg = c.page_config;
  // No config -> generic 4-tile fallback
  if(!cfg || !cfg.kpis || !cfg.kpis.length){
    var oee = (m.av!=null&&m.pf!=null)?m.av*m.pf:null;
    return [
      kpiTile('Tons', fK(m.tn), 'selected period', cColor, 75, 'sk'),
      kpiTile('Availability', fP(m.av), 'run / scheduled', s.accent, m.av?m.av*100:0, 'sk'),
      kpiTile('Performance', fP(m.pf), c.label==='Sand Plant' ? 'efficiency' : 'actual / rated TPH', '#107c10', m.pf?m.pf*100:0, 'sp'),
      kpiTile('OEE', fP(oee), 'Avail \u00d7 Perf', '#e66c37', oee?oee*100:0, 'so')
    ];
  }
  return cfg.kpis.map(function(k){
    var color = (k.color==='@accent') ? s.accent : (k.color || s.accent);
    var cls = k.cls || '';
    var val='\u2014', meta='', pct=0;
    if(k.calc==='tph'){
      val = (m.tph!=null) ? (Math.round(m.tph)+' TPH') : '\u2014';
      meta = k.target ? ('vs '+k.target.toLocaleString()+' TPH target') : 'tons / run hrs';
      pct = (m.tph!=null && k.target) ? Math.min(100, (m.tph/k.target)*100) : (m.tph!=null?70:0);
    } else if(k.calc==='tons'){
      val = fK(m.tn); meta = 'selected period'; pct = 75;
    } else if(k.calc==='est_tons'){
      val = fK(m.et); meta = 'estimated production'; pct = 60;
    } else if(k.calc==='av'){
      val = fP(m.av); meta = 'run / scheduled'; pct = m.av?m.av*100:0;
    } else if(k.calc==='pf'){
      val = fP(m.pf); meta = c.label==='Sand Plant' ? 'efficiency' : 'actual / rated TPH'; pct = m.pf?m.pf*100:0;
    } else if(k.calc==='oee'){
      var oe = (m.av!=null&&m.pf!=null)?m.av*m.pf:null;
      val = fP(oe); meta = 'Avail \u00d7 Perf'; pct = oe?oe*100:0;
    } else if(k.calc==='prod_tons'){
      var pa = productAgg(m._rows, k.product);
      val = fK(pa.tn); meta = k.goal_pct ? ('Goal: '+Math.round(k.goal_pct*100)+'% of output') : (k.product+' tons'); pct = 70;
    } else if(k.calc==='prod_yield'){
      var py = productAgg(m._rows, k.product);
      val = fP(py.yld); meta = k.product+' yield'; pct = py.yld?py.yld*100:0;
    }
    return kpiTile(k.l, val, meta, color, pct, cls);
  });
}

// ---- Sidebar renders ----
function renderGrainButtons(){
  document.getElementById('gbr').innerHTML = GRAINS.map(function(g){return '<button class="gb'+(ST.grain===g?' on':'')+'" onclick="setGrain(\''+g+'\')">'+g+'</button>';}).join('');
}
function renderYearSelect(){
  var el = document.getElementById('yr-sl');
  var yrs = allYearsForSite(ST.site);
  el.innerHTML = '<option value="All">All</option>' + yrs.map(function(y){return '<option value="'+y+'"'+(ST.yr===y?' selected':'')+'>'+y+'</option>';}).join('');
  el.onchange = function(){ ST.yr = el.value==='All'?'All':parseInt(el.value); ST.pidx=0; update(); };
}
function renderQtrSelect(){
  var el = document.getElementById('qt-sl');
  el.innerHTML = ['All','Q1','Q2','Q3','Q4'].map(function(q){return '<option value="'+q+'"'+(ST.qt===q?' selected':'')+'>'+q+'</option>';}).join('');
  el.onchange = function(){ ST.qt = el.value; update(); };
}
function renderMoSelect(){
  var el = document.getElementById('mo-sl');
  el.innerHTML = '<option value="All">All</option>' + MO.map(function(m){return '<option value="'+m+'"'+(ST.mo===m?' selected':'')+'>'+m+'</option>';}).join('');
  el.onchange = function(){ ST.mo = el.value; update(); };
}
function renderCircSelect(){
  var s = curSite();
  document.getElementById('circ-sl').innerHTML = s.circuit_order_active.map(function(k){
    var lbl = s.circuits[k].label;
    return '<div class="si" onclick="setCirc(\''+k+'\')"><div class="srb'+(ST.circ===k?' on':'')+'"></div><div class="sl">'+lbl+'</div></div>';
  }).join('');
}
function renderChips(){
  var chips = [];
  chips.push('<span class="chip site">'+ST.site+'</span>');
  chips.push('<span class="chip">'+ST.grain+'</span>');
  if(ST.yr!=='All') chips.push('<span class="chip">Year: '+ST.yr+'</span>');
  if(ST.qt!=='All') chips.push('<span class="chip">'+ST.qt+'</span>');
  if(ST.mo!=='All') chips.push('<span class="chip">'+ST.mo+'</span>');
  var s = curSite();
  if(s.circuits[ST.page]){ chips.push('<span class="chip">'+s.circuits[ST.page].label+'</span>'); }
  document.getElementById('chips').innerHTML = chips.join('');
}
function renderPtabs(){
  var pages = pagesForSite(ST.site);
  document.getElementById('ptabs').innerHTML = pages.map(function(p){return '<button class="pt'+(ST.page===p.k?' on':'')+'" onclick="setPage(\''+p.k+'\')">'+p.label+'</button>';}).join('');
  // Bookmark navigator = same page list, top-of-main, wireframe pattern
  document.getElementById('bnv').innerHTML = pages.map(function(p){return '<button class="bb'+(ST.page===p.k?' on':'')+'" onclick="setPage(\''+p.k+'\')">'+p.label+'</button>';}).join('');
}
function renderSiteToggle(){
  document.getElementById('site-tog').innerHTML = DATA.site_order.map(function(code){
    var s = SITES_BY_CODE[code];
    var on = (ST.site===code);
    // wireframe pattern: colored text + colored 3px bottom border when active
    var style = on ? ('color:'+s.accent+';border-bottom-color:'+s.accent+';') : '';
    return '<button class="stbtn'+(on?' on':'')+'" onclick="setSite(\''+code+'\')" style="'+style+'" title="'+escH(s.name)+'">'+code+'</button>';
  }).join('');
}
function renderSiteStrip(){
  var s = curSite();
  document.getElementById('site-strip').innerHTML =
    '<div><div class="site-id">'+s.code+'</div><div class="site-nm">'+s.name+'</div></div>';
}

function setPage(k){
  ST.page = k;
  var s = curSite();
  if(s.circuits[k]) ST.circ = k;
  update();
}
function setGrain(g){ST.grain=g; ST.pidx=0; update();}
function setCirc(k){
  ST.circ = k;
  var s = curSite();
  if(s.circuits[ST.page]) ST.page = k;
  update();
}
function setSite(code){
  if(ST.site===code) return;
  ST.site = code;
  var s = curSite();
  var yrs = allYearsForSite(code);
  var cur = new Date(DATA.as_of).getFullYear();
  ST.yr = yrs.indexOf(cur) >= 0 ? cur : (yrs[yrs.length-1] || 'All');
  ST.circ = s.circuit_order_active[0];
  // If current page is circuit-specific from previous site, reset to overview
  if(ST.page!=='overview' && ST.page!=='oee' && ST.page!=='dq' && !s.circuits[ST.page]){
    ST.page = 'overview';
  }
  applyAccent();
  update();
}
function stepP(d){ST.pidx=(ST.pidx||0)+d; update();}

function applyAccent(){
  var acc = curSite().accent;
  document.documentElement.style.setProperty('--accent', acc);
  document.documentElement.style.setProperty('--accent-tint', hexToRgba(acc, 0.12));
}

// ---- Stale / warn bar ----
function renderWbar(){
  var s = curSite();
  var bar = '';
  if(s.stale_file){
    bar = '<div class="warn-bar"><b>&#9888;</b> '+s.code+' workbook last modified '+s.file_mtime.replace('T',' ')+' &mdash; older than '+DATA.stale_hours+'h. Data may be stale.</div>';
  }
  // Also highlight active circuits with no data in last 14 days
  var today = new Date(DATA.as_of+'T00:00:00');
  var cutoff = new Date(today.getTime()-14*86400000);
  var stalecs = s.circuit_order_active.filter(function(k){
    var rs = s.circuits[k].rows.filter(function(r){return r._p && r._d>=cutoff;});
    return rs.length===0;
  });
  if(stalecs.length){
    bar += '<div class="warn-bar err"><b>&#9888;</b> No data in last 14 days on: '+stalecs.map(function(k){return s.circuits[k].label;}).join(', ')+'</div>';
  }
  document.getElementById('wbar').innerHTML = bar;
}

// ============ PAGE: Overview ============
function pageOverview(){
  var s = curSite();
  var order = s.circuit_order_active;
  var totTn=0, totRt=0, totSr=0, avAcc=[], pfAcc=[];
  order.forEach(function(k){
    var m = circMetrics(k);
    if(m.tn!=null) totTn += m.tn;
    if(m.rt!=null) totRt += m.rt;
    if(m.sr!=null) totSr += m.sr;
    if(m.av!=null) avAcc.push(m.av);
    if(m.pf!=null) pfAcc.push(m.pf);
  });
  var avAll = avg(avAcc), pfAll = avg(pfAcc);
  var oee = (avAll!=null&&pfAll!=null)?avAll*pfAll:null;
  var kpis = [
    kpiTile('Site Tons', fK(totTn), 'sum across circuits', s.accent, 80, 'sk'),
    kpiTile('Site Availability', fP(avAll), 'avg across circuits', s.accent, avAll?avAll*100:0, 'sk'),
    kpiTile('Site Performance', fP(pfAll), 'avg across circuits', '#107c10', pfAll?pfAll*100:0, 'sp'),
    kpiTile('Site OEE', fP(oee), 'Availability \u00d7 Performance', '#e66c37', oee?oee*100:0, 'so')
  ];
  document.getElementById('kg-w').innerHTML = '<div class="kg">'+kpis.join('')+'</div>';

  document.getElementById('r2').innerHTML =
    vb('ov1','Tons by circuit &mdash; stacked', 'Site production composition', legendByCirc(order))+
    vb('ov2','Availability \u00d7 Performance by circuit', 'Selected-period averages', [[s.accent,'Avail'],['#107c10','Perf']]);
  document.getElementById('r3').innerHTML =
    vb('ov3','Tons share by circuit &mdash; donut', 'Period totals')+
    vb('ov4','Site tons trend', 'Sum across circuits', [[s.accent,'Tons']]);

  var series = order.map(function(k){
    var agg = aggregate(filterRows(s.circuits[k].rows), 'tn', 'sum');
    return {key:k, agg:agg};
  });
  var base = series.reduce(function(a,b){return b.agg.labels.length>a.agg.labels.length?b:a;}, series[0]);
  var labels = base.agg.labels;
  var labelKeys = base.agg.keys;
  mCh('ov1','bar',labels, series.map(function(sr){
    var byKey={}; sr.agg.keys.forEach(function(k,i){byKey[k]=sr.agg.values[i];});
    return {label:s.circuits[sr.key].label, data:labelKeys.map(function(k){return byKey[k]||0;}), backgroundColor:circColor(sr.key), borderRadius:2};
  }), tk(), true);

  var cn = order.map(function(k){return s.circuits[k].label;});
  var avs = order.map(function(k){return circMetrics(k).av;});
  var pfs = order.map(function(k){return circMetrics(k).pf;});
  mCh('ov2','bar',cn,[
    {label:'Avail',data:avs,backgroundColor:hexToRgba(s.accent,0.75),borderRadius:2},
    {label:'Perf',data:pfs,backgroundColor:'rgba(16,124,16,0.75)',borderRadius:2}
  ],pf());

  var tns = order.map(function(k){return circMetrics(k).tn||0;});
  mCh('ov3','doughnut',cn,[{data:tns, backgroundColor:order.map(function(k){return circColor(k);}), borderWidth:1, borderColor:'#fff', hoverOffset:4}], null, false);
  if(CH['ov3']){CH['ov3'].options.plugins.legend={display:true,position:'right',labels:{font:{family:'Segoe UI',size:9},boxWidth:9,padding:4,color:'#323130'}};CH['ov3'].update();}

  var sumVals = labels.map(function(_,i){var t=0; series.forEach(function(sr){var v=sr.agg.values[sr.agg.keys.indexOf(labelKeys[i])]; if(v!=null) t+=v;}); return t;});
  mCh('ov4','line',labels,[{label:'Tons',data:sumVals,borderColor:s.accent,backgroundColor:hexToRgba(s.accent,0.07),fill:true,tension:0.35,pointRadius:3,borderWidth:2}], tk());
}

// ============ PAGE: OEE ============
function pageOEE(){
  var s = curSite();
  var order = s.circuit_order_active;
  var av={}, pfm={};
  order.forEach(function(k){ var m=circMetrics(k); av[k]=m.av; pfm[k]=m.pf; });
  var avAll=avg(Object.values(av).filter(function(x){return x!=null;}));
  var pfAll=avg(Object.values(pfm).filter(function(x){return x!=null;}));
  var maxav=-1,maxk=null; order.forEach(function(k){if(av[k]!=null&&av[k]>maxav){maxav=av[k];maxk=k;}});
  var maxpf=-1,maxp=null; order.forEach(function(k){if(pfm[k]!=null&&pfm[k]>maxpf){maxpf=pfm[k];maxp=k;}});

  var kpis=[
    kpiTile('Site Availability', fP(avAll), 'avg across circuits', s.accent, avAll?avAll*100:0, 'sk'),
    kpiTile('Site Performance', fP(pfAll), 'avg across circuits', '#107c10', pfAll?pfAll*100:0, 'sp'),
    kpiTile('Best Avail.', maxk?fP(maxav)+' &mdash; '+s.circuits[maxk].label:'\u2014', 'leading circuit', s.accent, maxav?maxav*100:0,''),
    kpiTile('Best Perf.', maxp?fP(maxpf)+' &mdash; '+s.circuits[maxp].label:'\u2014', 'leading circuit', '#107c10', maxpf?maxpf*100:0, '')
  ];
  document.getElementById('kg-w').innerHTML='<div class="kg">'+kpis.join('')+'</div>';

  // Circuit subsets — only circuits that actually carry the metric get a line on oe3/oe4
  var availOrder = order.filter(function(k){ return s.circuits[k].has_availability !== false; });
  var perfOrder  = order.filter(function(k){ return s.circuits[k].has_performance  !== false; });

  document.getElementById('r2').innerHTML=
    vb('oe1','Availability vs Performance by circuit','Selected-period averages',[[s.accent,'Avail'],['#107c10','Perf']])+
    vb('oe2','Run Time vs Scheduled &mdash; by circuit','Selected period',[[s.accent,'Run hrs'],['#8a8886','Scheduled hrs']]);
  document.getElementById('r3').innerHTML=
    vb('oe3','Availability over time &mdash; all circuits', null, availOrder.map(function(k){return [circColor(k),s.circuits[k].label];}))+
    vb('oe4','Performance over time &mdash; all circuits', null, perfOrder.map(function(k){return [circColor(k),s.circuits[k].label];}));

  var cn = order.map(function(k){return s.circuits[k].label;});
  var avs = order.map(function(k){return av[k];});
  var pfs = order.map(function(k){return pfm[k];});
  mCh('oe1','bar',cn,[
    {label:'Avail',data:avs,backgroundColor:hexToRgba(s.accent,0.75),borderRadius:2},
    {label:'Perf',data:pfs,backgroundColor:'rgba(16,124,16,0.75)',borderRadius:2}
  ],pf());

  var rts = order.map(function(k){return circMetrics(k).rt;});
  var srs = order.map(function(k){return circMetrics(k).sr;});
  mCh('oe2','bar',cn,[
    {label:'Run',data:rts,backgroundColor:hexToRgba(s.accent,0.75),borderRadius:2},
    {label:'Sched',data:srs,backgroundColor:'rgba(138,136,134,0.75)',borderRadius:2}
  ], function(v){return Math.round(v);});

  // oe3 — Availability over time (skip circuits without availability data)
  if(availOrder.length){
    var seriesAv = availOrder.map(function(k){return {key:k, agg:aggregate(filterRows(s.circuits[k].rows),'av','avg')};});
    var baseA = seriesAv.reduce(function(a,b){return b.agg.labels.length>a.agg.labels.length?b:a;}, seriesAv[0]);
    mCh('oe3','line',baseA.agg.labels, seriesAv.map(function(sr){
      var byKey={}; sr.agg.keys.forEach(function(k,i){byKey[k]=sr.agg.values[i];});
      return {label:s.circuits[sr.key].label, data:baseA.agg.keys.map(function(k){return byKey[k]==null?null:byKey[k];}), borderColor:circColor(sr.key), backgroundColor:'transparent', fill:false, tension:0.35, pointRadius:2, borderWidth:1.5, spanGaps:true};
    }), pf());
  }

  // oe4 — Performance over time (skip circuits without performance data)
  if(perfOrder.length){
    var seriesPf = perfOrder.map(function(k){return {key:k, agg:aggregate(filterRows(s.circuits[k].rows),'pf','avg')};});
    var baseP = seriesPf.reduce(function(a,b){return b.agg.labels.length>a.agg.labels.length?b:a;}, seriesPf[0]);
    mCh('oe4','line',baseP.agg.labels, seriesPf.map(function(sr){
      var byKey={}; sr.agg.keys.forEach(function(k,i){byKey[k]=sr.agg.values[i];});
      return {label:s.circuits[sr.key].label, data:baseP.agg.keys.map(function(k){return byKey[k]==null?null:byKey[k];}), borderColor:circColor(sr.key), backgroundColor:'transparent', fill:false, tension:0.35, pointRadius:2, borderWidth:1.5, spanGaps:true};
    }), pf());
  }
}

// ============ PAGE: per-circuit ============
function pageCircuit(){
  var s = curSite();
  var k = ST.page;
  var c = s.circuits[k];
  if(!c){ ST.page='overview'; return pageOverview(); }
  var m = circMetrics(k);
  var cColor = circColor(k);

  // Prefer page_config.kpis from site registry; fall back to generic 4-tile set
  var kpis = configuredKpis(c, m, s, cColor);
  document.getElementById('kg-w').innerHTML='<div class="kg">'+kpis.join('')+'</div>';

  var rows = filterRows(c.rows);
  var aggTn = aggregate(rows, 'tn', 'sum');
  var aggAv = aggregate(rows, 'av', 'avg');
  var aggPf = aggregate(rows, 'pf', 'avg');
  var aggRt = aggregate(rows, 'rt', 'sum');
  var aggSr = aggregate(rows, 'sr', 'sum');
  // Derived TPH per grain: tons_sum / run_hrs_sum
  var aggTph = aggTn.labels.map(function(_,i){
    var t=aggTn.values[i], r=aggRt.values[i];
    return (t!=null&&r!=null&&r>0) ? t/r : null;
  });

  var tphTarget = c.page_config && c.page_config.tph_target ? c.page_config.tph_target : null;
  var primaryChartTitle = tphTarget ? ('Avg TPH &mdash; dashed '+tphTarget.toLocaleString()+' TPH target') : 'Tons over time';
  var primaryChartLegend = tphTarget ? [[s.accent,'Avg TPH'],['#f2c811','Target']] : [[cColor,'Tons']];

  var r2html =
    vb(k+'_c1', primaryChartTitle, c.label+' \u00b7 '+ST.grain.toLowerCase(), primaryChartLegend)+
    vb(k+'_c2','Availability \u00d7 Performance trend', null, [[s.accent,'Avail'],['#107c10','Perf']]);
  document.getElementById('r2').innerHTML = r2html;
  // c4 title + legend (prod-mix w/ optional goal markers)
  var c4HasGoals = false;
  if(c.has_products){
    if(s.goals && s.goals.products){
      c4HasGoals = c.product_names.some(function(p){return s.goals.products[p];});
    }
    if(!c4HasGoals && c.page_config && c.page_config.kpis){
      c4HasGoals = c.page_config.kpis.some(function(kp){return kp.goal_pct && (kp.calc==='prod_tons'||kp.calc==='prod_yield');});
    }
  }
  var c4Title = c.has_trucks ? 'Haul fleet activity (load counts)' : (c.has_products ? ('Product mix &mdash; tons'+(c4HasGoals?' (goal markers)':'')) : 'Daily tons');
  var c4Legend = c.has_trucks ? c.truck_labels.map(function(t,i){return [CIRC_PALETTES[s.code][i%CIRC_PALETTES[s.code].length], t];}) : (c.has_products ? c.product_names.map(function(p,i){return [CIRC_PALETTES[s.code][i%CIRC_PALETTES[s.code].length], p];}) : [[cColor,'Tons']]);
  if(c4HasGoals) c4Legend.push(['#f2c811','Product goal']);
  document.getElementById('r3').innerHTML =
    vb(k+'_c3','Run Time vs Scheduled', null, [[s.accent,'Run'],['#8a8886','Scheduled']])+
    vb(k+'_c4', c4Title, null, c4Legend);

  if(tphTarget){
    // Line chart: TPH with dashed target overlay
    mCh(k+'_c1','line', aggTn.labels, [
      {label:'TPH', data:aggTph, borderColor:s.accent, backgroundColor:hexToRgba(s.accent,0.07), fill:true, tension:0.35, pointRadius:3, borderWidth:2, spanGaps:true},
      {label:'Target', data:aggTn.labels.map(function(){return tphTarget;}), borderColor:'#f2c811', borderWidth:1.5, borderDash:[5,4], pointRadius:0, fill:false}
    ], function(v){return Math.round(v);});
  } else {
    mCh(k+'_c1','bar', aggTn.labels, [{label:'Tons', data:aggTn.values, backgroundColor:cColor, borderRadius:2}], tk());
  }
  mCh(k+'_c2','line',aggAv.labels,[
    {label:'Avail',data:aggAv.values,borderColor:s.accent,backgroundColor:hexToRgba(s.accent,0.07),fill:true,tension:0.35,pointRadius:2,borderWidth:1.5,spanGaps:true},
    {label:'Perf',data:aggPf.values,borderColor:'#107c10',fill:false,tension:0.35,pointRadius:2,borderWidth:1.5,spanGaps:true}
  ], pf());
  mCh(k+'_c3','bar',aggRt.labels,[
    {label:'Run',data:aggRt.values,backgroundColor:hexToRgba(s.accent,0.75),borderRadius:2},
    {label:'Sched',data:aggSr.values,backgroundColor:'rgba(138,136,134,0.75)',borderRadius:2}
  ], function(v){return Math.round(v);});

  // c4: trucks OR products OR plain tons
  if(c.has_trucks){
    var keys=[], byKey={};
    rows.forEach(function(r){
      var key;
      if(ST.grain==='Daily') key=r.d;
      else if(ST.grain==='Weekly') key=r._y+'-W'+String(isoWeek(r._d)).padStart(2,'0');
      else if(ST.grain==='Monthly') key=r._y+'-'+String(r._m+1).padStart(2,'0');
      else if(ST.grain==='Quarterly') key=r._y+'-'+r._q;
      else key=String(r._y);
      if(byKey[key]===undefined){byKey[key]={}; c.truck_labels.forEach(function(t){byKey[key][t]=0;}); keys.push(key);}
      if(r.tk){c.truck_labels.forEach(function(t){byKey[key][t]+=(r.tk[t]||0);});}
    });
    keys.sort();
    var labels = keys.map(function(kk){
      if(ST.grain==='Monthly'){var p=kk.split('-');return MO[parseInt(p[1])-1]+(ST.yr==='All'?' '+p[0].slice(2):'');}
      if(ST.grain==='Quarterly'){var p=kk.split('-');return p[1]+(ST.yr==='All'?' '+p[0].slice(2):'');}
      if(ST.grain==='Weekly'){var p=kk.split('-W');return 'W'+p[1]+(ST.yr==='All'?" '"+p[0].slice(2):'');}
      if(ST.grain==='Daily'){var d=new Date(kk+'T00:00:00'); return (d.getMonth()+1)+'/'+d.getDate();}
      return kk;
    });
    var pal = CIRC_PALETTES[s.code];
    var ds = c.truck_labels.map(function(t,i){return {label:t, data:keys.map(function(kk){return byKey[kk][t];}), backgroundColor:pal[i%pal.length], borderRadius:2};});
    mCh(k+'_c4','bar',labels, ds, function(v){return Math.round(v);}, true);
  } else if(c.has_products){
    // Aggregate product tons by grain key
    var keys=[], byKey={};
    rows.forEach(function(r){
      var key;
      if(ST.grain==='Daily') key=r.d;
      else if(ST.grain==='Weekly') key=r._y+'-W'+String(isoWeek(r._d)).padStart(2,'0');
      else if(ST.grain==='Monthly') key=r._y+'-'+String(r._m+1).padStart(2,'0');
      else if(ST.grain==='Quarterly') key=r._y+'-'+r._q;
      else key=String(r._y);
      if(byKey[key]===undefined){byKey[key]={}; c.product_names.forEach(function(p){byKey[key][p]=0;}); keys.push(key);}
      if(r.pm){c.product_names.forEach(function(p){var v=r.pm[p]&&r.pm[p].tn; if(v!=null) byKey[key][p]+=v;});}
    });
    keys.sort();
    var labels = keys.map(function(kk){
      if(ST.grain==='Monthly'){var p=kk.split('-');return MO[parseInt(p[1])-1]+(ST.yr==='All'?' '+p[0].slice(2):'');}
      if(ST.grain==='Quarterly'){var p=kk.split('-');return p[1]+(ST.yr==='All'?' '+p[0].slice(2):'');}
      if(ST.grain==='Weekly'){var p=kk.split('-W');return 'W'+p[1]+(ST.yr==='All'?" '"+p[0].slice(2):'');}
      if(ST.grain==='Daily'){var d=new Date(kk+'T00:00:00'); return (d.getMonth()+1)+'/'+d.getDate();}
      return kk;
    });
    var pal = CIRC_PALETTES[s.code];
    var ds = c.product_names.map(function(p,i){return {label:p, data:keys.map(function(kk){return byKey[kk][p];}), backgroundColor:pal[i%pal.length], borderRadius:2};});
    // Wireframe goal-marker overlay: yellow crossRot points at goal_pct of bucket total for each product with a goal.
    // Source of truth: site.goals.products (from Goals sheet), falling back to page_config.kpis.
    var goalsMap = {};
    if(s.goals && s.goals.products){
      Object.keys(s.goals.products).forEach(function(p){
        if(c.product_names.indexOf(p) >= 0) goalsMap[p] = s.goals.products[p].pct;
      });
    }
    if(c.page_config && c.page_config.kpis){
      c.page_config.kpis.forEach(function(kp){
        if((kp.calc==='prod_tons' || kp.calc==='prod_yield') && kp.goal_pct && kp.product && goalsMap[kp.product]===undefined){
          goalsMap[kp.product] = kp.goal_pct;
        }
      });
    }
    var goalProducts = Object.keys(goalsMap);
    if(goalProducts.length){
      var bucketTotals = keys.map(function(kk){var t=0; c.product_names.forEach(function(p){t+=(byKey[kk][p]||0);}); return t;});
      goalProducts.forEach(function(p){
        var gpct = goalsMap[p];
        var goalData = bucketTotals.map(function(t){return t>0 ? t*gpct : null;});
        ds.push({
          type:'line', label:p+' goal ('+Math.round(gpct*100)+'%)',
          data:goalData, borderColor:'#f2c811', backgroundColor:'#f2c811',
          showLine:false, pointStyle:'crossRot', pointRadius:7, pointHoverRadius:9,
          pointBorderWidth:2, borderWidth:0, fill:false, stack:'goal-'+p, yAxisID:'y'
        });
      });
    }
    mCh(k+'_c4','bar',labels, ds, function(v){return Math.round(v);}, true);
  } else {
    mCh(k+'_c4','bar',aggTn.labels,[{label:'Tons',data:aggTn.values,backgroundColor:cColor,borderRadius:2}], tk());
  }
}

// ============ PAGE: Data Quality ============
function pageDQ(){
  var s = curSite();
  var today = new Date(DATA.as_of+'T00:00:00');
  // Build KPI tiles showing last-updated age for every circuit (active + stale)
  var all = s.circuit_order_active.concat(s.circuit_order_stale);
  var kpis = all.map(function(k){
    var c = s.circuits[k];
    var rs = c.rows;
    var last=null; for(var i=rs.length-1;i>=0;i--){if(rs[i]._p){last=rs[i]._d;break;}}
    var age = last?Math.round((today-last)/86400000):null;
    var cls = c.stale ? 'off' : (age==null?'off':(age<=2?'sk':age<=7?'so':'sr'));
    var color = c.stale ? '#8a8886' : (age==null?'#8a8886':(age<=2?'#107c10':age<=7?'#e66c37':'#a4262c'));
    var label = c.label + (c.stale ? ' (stale)' : '');
    var meta = age==null?'no data':age+' d old';
    return kpiTile(label, meta, last?'last '+last.toISOString().slice(0,10):'\u2014', color, age==null?0:Math.max(0,100-age*3), cls);
  });
  document.getElementById('kg-w').innerHTML='<div class="kg">'+kpis.join('')+'</div>';

  document.getElementById('r2').innerHTML = vb('dq1','Populated days per circuit (last 30)','Daily data capture rate',[[s.accent,'Days w/ data']]) +
    vb('dq2','Recent downtime notes','Last 10 comments across circuits','');
  document.getElementById('r3').innerHTML = vb('dq3','Stale circuits &mdash; data coverage','Circuits with no recent entries',[['#a4262c','Days since last entry']]) +
    vb('dq4','Product mix &mdash; site totals','Selected-period product tons (where tracked)','');

  var cutoff = new Date(today.getTime()-29*86400000);
  var activeCn = s.circuit_order_active.map(function(k){return s.circuits[k].label;});
  var pop = s.circuit_order_active.map(function(k){return s.circuits[k].rows.filter(function(r){return r._p && r._d>=cutoff && r._d<=today;}).length;});
  mCh('dq1','bar',activeCn,[{label:'Days',data:pop,backgroundColor:hexToRgba(s.accent,0.75),borderRadius:2}], function(v){return Math.round(v);});

  // dq2: downtime notes
  var rows=[];
  s.circuit_order_active.concat(s.circuit_order_stale).forEach(function(k){
    var c = s.circuits[k];
    if(!c.has_downtime_note) return;
    c.rows.forEach(function(r){ if(r.dt) rows.push({d:r.d, lbl:c.label, txt:r.dt}); });
  });
  rows.sort(function(a,b){return b.d.localeCompare(a.d);});
  rows = rows.slice(0,10);
  var tbl = '<table class="mx"><thead><tr><th>Date</th><th>Circuit</th><th>Comment</th></tr></thead><tbody>'+
    (rows.length?rows.map(function(r){return '<tr><td>'+r.d+'</td><td>'+escH(r.lbl)+'</td><td style="white-space:normal;max-width:380px">'+escH(r.txt)+'</td></tr>';}).join(''):'<tr><td colspan="3">No downtime comments found.</td></tr>')+
    '</tbody></table>';
  document.getElementById('dq2').querySelector('.vb').innerHTML = '<div class="mxw">'+tbl+'</div>';

  // dq3: stale circuits (days since last entry)
  var staleCn = s.circuit_order_stale.map(function(k){return s.circuits[k].label;});
  var staleAges = s.circuit_order_stale.map(function(k){
    var rs=s.circuits[k].rows; var last=null;
    for(var i=rs.length-1;i>=0;i--){if(rs[i]._p){last=rs[i]._d;break;}}
    return last?Math.round((today-last)/86400000):null;
  });
  if(staleCn.length){
    mCh('dq3','bar',staleCn,[{label:'Days',data:staleAges,backgroundColor:'rgba(164,38,44,0.75)',borderRadius:2}], function(v){return Math.round(v);});
  } else {
    document.getElementById('dq3').querySelector('.vb').innerHTML = '<div class="nd">No stale circuits on this site.</div>';
  }

  // dq4: product totals across circuits that have product matrices
  var prodRows=[];
  s.circuit_order_active.forEach(function(k){
    var c = s.circuits[k];
    if(!c.has_products) return;
    var rs = filterRows(c.rows);
    c.product_names.forEach(function(p){
      var t = sum(rs.map(function(r){return r.pm && r.pm[p] ? r.pm[p].tn : null;}));
      prodRows.push({circ:c.label, product:p, tons:t||0});
    });
  });
  if(prodRows.length){
    prodRows.sort(function(a,b){return b.tons-a.tons;});
    var tbl2 = '<table class="mx"><thead><tr><th>Circuit</th><th>Product</th><th>Tons</th></tr></thead><tbody>'+
      prodRows.slice(0,25).map(function(r){return '<tr><td>'+escH(r.circ)+'</td><td>'+escH(r.product)+'</td><td>'+fK(r.tons)+'</td></tr>';}).join('')+
      '</tbody></table>';
    document.getElementById('dq4').querySelector('.vb').innerHTML = '<div class="mxw">'+tbl2+'</div>';
  } else {
    document.getElementById('dq4').querySelector('.vb').innerHTML = '<div class="nd">No product-matrix circuits on this site.</div>';
  }

  /* === CONFIG HEALTH === */
  renderConfigHealth(s);
}

function renderConfigHealth(s){
  /* Roll up per-circuit config_health tallies into site-level summary + warning table.
     Source fields used:
       s.config_health[ckey] = { label, sheet, header_row, per_metric, status_tally, warnings, sanity }
       s.unregistered_sheets = [{sheet, detected_metrics, ...}]
       s.site_warnings = [str]
       s.goals.resolution = { header_row, by, warning }
     DATA.schema_drift (global) = [{site, kind, circuit?, metric?, from, to, sheet?}]
  */
  var mxt = document.getElementById('mxt');
  var mxb = document.getElementById('mxb');
  if(!mxt || !mxb) return;
  mxt.textContent = 'Config Health — header resolution, data sanity, and schema drift';

  var ch = s.config_health || {};
  var circOrder = s.circuit_order_active.concat(s.circuit_order_stale);
  var siteTally = {confirmed:0, relocated:0, unverified:0, detected:0, not_configured:0};
  var warnCount = 0, oorTotal = 0, futureTotal = 0;
  circOrder.forEach(function(k){
    var b = ch[k]; if(!b) return;
    Object.keys(b.status_tally||{}).forEach(function(t){ siteTally[t] += (b.status_tally[t]||0); });
    warnCount += (b.warnings||[]).length;
    if(b.sanity){
      var oor = b.sanity.oor_counts || {};
      Object.keys(oor).forEach(function(k2){ oorTotal += (oor[k2]||0); });
      futureTotal += ((b.sanity.suspect_future_dates||[]).length);
    }
  });

  function badge(label, n, color){
    var bg = (n>0 && color) ? color : '#edebe9';
    var fg = (n>0 && color) ? '#fff' : '#605e5c';
    return '<span style="display:inline-block;padding:2px 8px;border-radius:2px;margin-right:6px;font-size:11px;font-weight:700;background:'+bg+';color:'+fg+'">'+escH(label)+': '+n+'</span>';
  }

  var html = '';
  html += '<div style="padding:10px 12px 6px 12px;font-size:11px;color:#605e5c;">';
  html += '<div style="margin-bottom:6px;font-weight:700;color:#201f1e;font-size:12px">Column resolution across '+circOrder.length+' circuits</div>';
  html += badge('Confirmed',   siteTally.confirmed,     '#107c10');
  html += badge('Relocated',   siteTally.relocated,     '#e66c37');
  html += badge('Unverified',  siteTally.unverified,    '#e66c37');
  html += badge('Detected (new)', siteTally.detected,    '#0078d4');
  html += badge('Not configured', siteTally.not_configured, null);
  html += badge('Total warnings', warnCount, warnCount>0 ? '#e66c37' : null);
  html += badge('OOR values dropped', oorTotal, oorTotal>0 ? '#e66c37' : null);
  html += badge('Future-dated rows', futureTotal, futureTotal>0 ? '#e66c37' : null);
  html += '</div>';

  /* Warnings table */
  var rows = [];
  circOrder.forEach(function(k){
    var b = ch[k]; if(!b) return;
    (b.warnings||[]).forEach(function(w){
      rows.push({circ:b.label, sheet:b.sheet, metric:w.metric, status:w.status, msg:w.message});
    });
  });
  if(rows.length){
    html += '<table class="mx" style="margin-top:6px"><thead><tr><th>Circuit</th><th>Sheet</th><th>Metric</th><th>Status</th><th>Message</th></tr></thead><tbody>';
    rows.forEach(function(r){
      var cBg = (r.status==='relocated') ? 'background:rgba(230,108,55,0.10)' :
                (r.status==='unverified') ? 'background:rgba(230,108,55,0.06)' :
                (r.status==='detected')   ? 'background:rgba(0,120,212,0.06)'  : '';
      html += '<tr style="'+cBg+'"><td>'+escH(r.circ)+'</td><td>'+escH(r.sheet)+'</td><td>'+escH(r.metric)+'</td><td><strong>'+escH(r.status)+'</strong></td><td style="white-space:normal;max-width:600px">'+escH(r.msg)+'</td></tr>';
    });
    html += '</tbody></table>';
  } else {
    html += '<div class="nd" style="margin-top:6px">All configured columns verified against sheet headers &mdash; no drift, no fallbacks.</div>';
  }

  /* Per-circuit sanity (OOR drops + future-dated row counts) */
  var sanRows = [];
  circOrder.forEach(function(k){
    var b = ch[k]; if(!b || !b.sanity) return;
    var oor = b.sanity.oor_counts || {};
    var oorSum = 0; Object.keys(oor).forEach(function(k2){ oorSum += (oor[k2]||0); });
    var futN = (b.sanity.suspect_future_dates||[]).length;
    if(oorSum === 0 && futN === 0) return;
    sanRows.push({
      circ: b.label, sheet: b.sheet,
      oor_av: oor.av||0, oor_pf: oor.pf||0, oor_tn: oor.tn||0, oor_et: oor.et||0,
      future: futN,
      sample: (b.sanity.oor_samples && (b.sanity.oor_samples.av[0] || b.sanity.oor_samples.pf[0] || b.sanity.oor_samples.tn[0])) || null,
      future_sample: (b.sanity.suspect_future_dates||[])[0] || null
    });
  });
  if(sanRows.length){
    html += '<div style="margin-top:12px;padding:0 12px 4px 12px;font-weight:700;color:#201f1e;font-size:12px">Data sanity exceptions</div>';
    html += '<table class="mx"><thead><tr><th>Circuit</th><th>Sheet</th><th>OOR Avail</th><th>OOR Perf</th><th>OOR Tons</th><th>OOR Est&nbsp;Tons</th><th>Future rows</th><th>Sample</th></tr></thead><tbody>';
    sanRows.forEach(function(r){
      var samp = r.sample ? ('row '+r.sample.row+'='+r.sample.value) : (r.future_sample ? ('row '+r.future_sample.row+' '+r.future_sample.date) : '');
      html += '<tr><td>'+escH(r.circ)+'</td><td>'+escH(r.sheet)+'</td><td>'+r.oor_av+'</td><td>'+r.oor_pf+'</td><td>'+r.oor_tn+'</td><td>'+r.oor_et+'</td><td>'+r.future+'</td><td>'+escH(samp)+'</td></tr>';
    });
    html += '</tbody></table>';
  }

  /* Goals resolution (RSQ only has goals) */
  var gres = s.goals && s.goals.resolution;
  if(gres && (gres.warning || gres.by)){
    var gc = (gres.warning) ? '#e66c37' : '#107c10';
    html += '<div style="margin-top:12px;padding:0 12px 4px 12px;font-weight:700;color:#201f1e;font-size:12px">Goals sheet resolution</div>';
    html += '<div style="padding:6px 12px;font-size:11px;">';
    html += '<span style="display:inline-block;padding:2px 8px;border-radius:2px;margin-right:6px;font-weight:700;background:'+gc+';color:#fff">'+escH((gres.by||'unknown'))+'</span>';
    if(gres.header_row) html += '<span style="color:#605e5c">header row '+gres.header_row+' &middot; </span>';
    var prodCount = Object.keys(s.goals.products||{}).length;
    html += '<span style="color:#605e5c">'+prodCount+' product(s) loaded</span>';
    if(gres.warning) html += '<div style="margin-top:4px;color:#a4262c">'+escH(gres.warning)+'</div>';
    html += '</div>';
  }

  /* Unregistered sheets */
  if((s.unregistered_sheets||[]).length){
    html += '<div style="margin-top:12px;padding:0 12px 4px 12px;font-weight:700;color:#201f1e;font-size:12px">Unregistered production-looking sheets</div>';
    html += '<table class="mx"><thead><tr><th>Sheet</th><th>Detected metrics</th><th>Size</th></tr></thead><tbody>';
    s.unregistered_sheets.forEach(function(u){
      html += '<tr><td>'+escH(u.sheet)+'</td><td>'+escH((u.detected_metrics||[]).join(', '))+'</td><td>'+u.max_row+' &times; '+u.max_col+'</td></tr>';
    });
    html += '</tbody></table>';
  }

  /* Schema drift vs previous snapshot */
  var drift = (DATA.schema_drift || []).filter(function(d){ return d.site === s.code; });
  if(drift.length){
    html += '<div style="margin-top:12px;padding:0 12px 4px 12px;font-weight:700;color:#201f1e;font-size:12px">Schema drift since last run ('+(DATA.schema_snapshot_prev||'n/a')+')</div>';
    html += '<table class="mx"><thead><tr><th>Kind</th><th>Circuit</th><th>Metric</th><th>From</th><th>To</th></tr></thead><tbody>';
    drift.forEach(function(d){
      html += '<tr><td><strong>'+escH(d.kind)+'</strong></td><td>'+escH(d.circuit||d.sheet||'')+'</td><td>'+escH(d.metric||'')+'</td><td>'+escH(String(d.from==null?'':d.from))+'</td><td>'+escH(String(d.to==null?'':d.to))+'</td></tr>';
    });
    html += '</tbody></table>';
  } else if(DATA.schema_snapshot_prev){
    html += '<div class="nd" style="margin-top:8px">No schema drift since '+escH(DATA.schema_snapshot_prev)+'.</div>';
  }

  /* Summary line per-circuit (headers at which row, how many metrics resolved) */
  html += '<div style="margin-top:12px;padding:0 12px 4px 12px;font-weight:700;color:#201f1e;font-size:12px">Per-circuit header resolution</div>';
  html += '<table class="mx"><thead><tr><th>Circuit</th><th>Sheet</th><th>Header row</th><th>Confirmed</th><th>Relocated</th><th>Unverified</th><th>Detected</th></tr></thead><tbody>';
  circOrder.forEach(function(k){
    var b = ch[k]; if(!b) return;
    var t = b.status_tally||{};
    html += '<tr><td>'+escH(b.label)+'</td><td>'+escH(b.sheet)+'</td><td>'+b.header_row+'</td>' +
            '<td>'+(t.confirmed||0)+'</td><td>'+(t.relocated||0)+'</td><td>'+(t.unverified||0)+'</td><td>'+(t.detected||0)+'</td></tr>';
  });
  html += '</tbody></table>';

  mxb.innerHTML = '<div class="mxw">'+html+'</div>';
}

// ---- visual box helper ----
function vb(id, title, note, legendPairs){
  var lg='';
  if(legendPairs && legendPairs.length){
    lg = '<div class="lg">'+legendPairs.map(function(p){return '<div class="li"><div class="ls" style="background:'+p[0]+'"></div>'+escH(p[1])+'</div>';}).join('')+'</div>';
  }
  return '<div class="vi" id="'+id+'"><div class="vh"><div class="vt">'+title+'</div></div>'+(note?'<div class="vn">'+note+'</div>':'')+lg+'<div class="vb"><div class="cw"><canvas id="cv_'+id+'"></canvas></div></div></div>';
}
function legendByCirc(order){var s=curSite(); return order.map(function(k){return [circColor(k), s.circuits[k].label];});}

// ---- chart helpers ----
function co(yf, stk){
  return {responsive:true, maintainAspectRatio:false, animation:{duration:200},
    plugins:{legend:{display:false}, tooltip:{titleFont:{family:'Segoe UI',size:10}, bodyFont:{family:'Segoe UI',size:10}, padding:5, backgroundColor:'#323130', titleColor:'#f3f2f1', bodyColor:'#c8c6c4'}},
    scales:{x:{grid:{color:'rgba(0,0,0,0.05)'}, ticks:{font:{family:'Segoe UI',size:9}, color:'#605e5c', maxRotation:45, autoSkip:true}, stacked:!!stk},
            y:{grid:{color:'rgba(0,0,0,0.06)'}, ticks:{font:{family:'Segoe UI',size:9}, color:'#605e5c', callback:yf||undefined}, stacked:!!stk}}};
}
function mCh(id, type, lbls, ds, yf, stk){ kc(id); var el=document.getElementById('cv_'+id); if(!el) el=document.getElementById(id); if(!el) return; CH[id]=new Chart(el,{type:type, data:{labels:lbls, datasets:ds}, options:co(yf,stk)}); }
function pf(){return function(v){return v!=null?(v*100).toFixed(0)+'%':'';}}
function tk(){return function(v){return v!=null&&v>=1000?(v/1000).toFixed(0)+'k':v!=null?Math.round(v)+'':'';}}

// ---- Matrix ----
function renderMatrix(){
  if(ST.page==='dq'){ /* DQ page renders Config Health into mxt/mxb itself */ return; }
  var s = curSite();
  var isCircuitPage = !!s.circuits[ST.page];
  var key = isCircuitPage ? ST.page : ST.circ;
  var c = s.circuits[key];
  document.getElementById('mxt').textContent = 'Matrix &mdash; '+c.label+' \u00b7 '+ST.grain+' aggregates';
  var rows = filterRows(c.rows);
  var agTn = aggregate(rows,'tn','sum');
  var agRt = aggregate(rows,'rt','sum');
  var agSr = aggregate(rows,'sr','sum');
  var agAv = aggregate(rows,'av','avg');
  var agPf = aggregate(rows,'pf','avg');
  var labels = agTn.labels;
  var totalSumTn = sum(agTn.values)||0;
  var avgAv = avg(agAv.values);
  var avgPf = avg(agPf.values);
  var totalRt = sum(agRt.values)||0;
  var totalSr = sum(agSr.values)||0;
  function row(name, vals, f, total){
    return '<tr><td>'+name+'</td>'+vals.map(function(v){return '<td>'+f(v)+'</td>';}).join('')+'<td>'+f(total)+'</td></tr>';
  }
  var h='<table class="mx"><thead><tr><th>Metric</th>'+labels.map(function(l){return '<th>'+escH(l)+'</th>';}).join('')+'<th>Total/Avg</th></tr></thead><tbody>';
  h += row('Tons', agTn.values, fK, sum(agTn.values));
  h += row('Run hrs', agRt.values, function(v){return fN(v,1);}, totalRt);
  h += row('Sched hrs', agSr.values, function(v){return fN(v,1);}, totalSr);
  if(c.has_availability) h += row('Availability', agAv.values, fP, avgAv);
  if(c.has_performance) h += row('Performance', agPf.values, fP, avgPf);

  // Product matrix addendum (if this circuit has one)
  if(c.has_products){
    h += '<tr><td colspan="'+(labels.length+2)+'" style="padding-top:8px;color:#605e5c;font-weight:700">Product mix (tons)</td></tr>';
    c.product_names.forEach(function(pname){
      // Aggregate this product's tons across grain keys
      var groups={}, order=[];
      rows.forEach(function(r){
        var k;
        if(ST.grain==='Daily') k=r.d;
        else if(ST.grain==='Weekly') k=r._y+'-W'+String(isoWeek(r._d)).padStart(2,'0');
        else if(ST.grain==='Monthly') k=r._y+'-'+String(r._m+1).padStart(2,'0');
        else if(ST.grain==='Quarterly') k=r._y+'-'+r._q;
        else k=String(r._y);
        if(groups[k]===undefined){groups[k]=0;order.push(k);}
        var v = r.pm && r.pm[pname] ? r.pm[pname].tn : null;
        if(v!=null) groups[k] += v;
      });
      order.sort();
      var vals = agTn.keys.map(function(k){return groups[k]!=null?groups[k]:null;});
      h += row(pname, vals, fK, sum(vals));
    });
  }
  h += '</tbody></table>';
  document.getElementById('mxb').innerHTML = h;
}

// ---- orchestrator ----
function update(){
  cr();
  renderSiteToggle();
  renderSiteStrip();
  renderGrainButtons();
  renderYearSelect();
  renderQtrSelect();
  renderMoSelect();
  renderCircSelect();
  renderChips();
  renderPtabs();
  renderWbar();
  var s = curSite();
  document.getElementById('src-file').textContent = s.file;
  document.getElementById('src-mtime').textContent = s.file_mtime.replace('T',' ');
  document.getElementById('src-asof').textContent = DATA.as_of;
  document.getElementById('refresh-lbl').textContent = 'Refreshed: '+DATA.generated_at;
  if(isRawGrain()){
    // KPI tiles still reflect the period filter (daily source is fine); chart grid becomes placeholder
    renderKpisForRawGrain();
    rawPlaceholder();
  } else {
    if(ST.page==='overview') pageOverview();
    else if(ST.page==='oee') pageOEE();
    else if(ST.page==='dq') pageDQ();
    else pageCircuit();
  }
  renderMatrix();
}

function renderKpisForRawGrain(){
  var s = curSite();
  var order = s.circuit_order_active;
  var totTn=0, avAcc=[], pfAcc=[];
  order.forEach(function(k){ var m=circMetrics(k); if(m.tn!=null) totTn+=m.tn; if(m.av!=null) avAcc.push(m.av); if(m.pf!=null) pfAcc.push(m.pf); });
  var avAll=avg(avAcc), pfAll=avg(pfAcc), oee=(avAll!=null&&pfAll!=null)?avAll*pfAll:null;
  var k=[
    kpiTile('Site Tons', fK(totTn), 'period total', s.accent, 75, 'sk'),
    kpiTile('Site Availability', fP(avAll), 'avg across circuits', s.accent, avAll?avAll*100:0, 'sk'),
    kpiTile('Site Performance', fP(pfAll), 'avg across circuits', '#107c10', pfAll?pfAll*100:0, 'sp'),
    kpiTile('Site OEE', fP(oee), 'Avail \u00d7 Perf', '#e66c37', oee?oee*100:0, 'so')
  ];
  document.getElementById('kg-w').innerHTML='<div class="kg">'+k.join('')+'</div>';
}

applyAccent();
update();
</script>
</body></html>
"""


# ============================================================================
# MAIN
# ============================================================================

def main():
    p = argparse.ArgumentParser()
    p.add_argument("--arq", required=True, help="path to Ardmore workbook")
    p.add_argument("--rsq", required=True, help="path to Richards Spur workbook")
    p.add_argument("--out", required=True, help="path to output HTML")
    p.add_argument("--chartjs", required=False,
                   default=str(Path(__file__).with_name("chart.umd.js")),
                   help="path to Chart.js UMD bundle (inlined into HTML)")
    p.add_argument("--tz", required=False, default=DEFAULT_TZ,
                   help="IANA timezone for the Refreshed label (default: %(default)s)")
    args = p.parse_args()

    chartjs_path = Path(args.chartjs)
    if not chartjs_path.exists():
        sys.stderr.write(f"warn: chart.umd.js not found at {chartjs_path}; "
                         f"dashboard will render no charts\n")

    build(Path(args.arq), Path(args.rsq), Path(args.out), chartjs_path, tz_name=args.tz)


if __name__ == "__main__":
    main()
# end
